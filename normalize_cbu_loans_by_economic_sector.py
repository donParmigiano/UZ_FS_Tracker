import argparse
import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUT = Path("data/master/cbu_loans_by_economic_sector_master.csv")
QA_OUT = Path("data/master/cbu_loans_by_economic_sector_parse_qa.csv")
FILE_TOKEN = "Information-on-commercial-bank-loans-by-economic-sector"


@dataclass
class ParsedRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    sector: str
    metric_type: str
    value: Optional[float]
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Normalize CBU report: Information on commercial bank loans by economic sector"
    )
    parser.add_argument(
        "--periods",
        type=str,
        default=None,
        help="Comma-separated YYYY_MM folders to parse (e.g. 2026_04,2024_04)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output files if they already exist",
    )
    return parser.parse_args()


def build_period_folder_list(periods_arg: Optional[str]) -> List[Path]:
    if periods_arg:
        folders: List[Path] = []
        for part in periods_arg.split(","):
            period = part.strip()
            if not period:
                continue
            folder = RAW_ROOT / period
            if folder.is_dir():
                folders.append(folder)
        return folders

    if not RAW_ROOT.exists():
        return []
    return sorted([p for p in RAW_ROOT.iterdir() if p.is_dir()])


def find_target_files(period_folders: Iterable[Path]) -> List[Tuple[str, Path]]:
    out: List[Tuple[str, Path]] = []
    for folder in period_folders:
        for file_path in sorted(folder.glob("*.xlsx")):
            if FILE_TOKEN in file_path.name:
                out.append((folder.name, file_path))
    return out


def extract_merged_header_values(ws) -> Dict[str, Optional[datetime]]:
    header_dates: Dict[str, Optional[datetime]] = {"B": None, "C": None, "D": None, "E": None}

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= 3 <= merged_range.max_row:
            if merged_range.min_col <= 2 <= merged_range.max_col and merged_range.max_col >= 3:
                value = ws.cell(row=3, column=merged_range.min_col).value
                header_dates["B"] = to_datetime(value)
                header_dates["C"] = to_datetime(value)
            if merged_range.min_col <= 4 <= merged_range.max_col and merged_range.max_col >= 5:
                value = ws.cell(row=3, column=merged_range.min_col).value
                header_dates["D"] = to_datetime(value)
                header_dates["E"] = to_datetime(value)

    for col in ["B", "C", "D", "E"]:
        if header_dates[col] is None:
            value = ws[f"{col}3"].value
            header_dates[col] = to_datetime(value)

    return header_dates


def to_datetime(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day)

    text = str(value).strip()
    patterns = [
        "%B %d, %Y",
        "%b %d, %Y",
        "%d.%m.%Y",
        "%Y-%m-%d",
    ]
    for pattern in patterns:
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def excel_value_to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def format_period_label(dt: datetime) -> str:
    return f"{dt.strftime('%B')} {dt.day}, {dt.year}"


def parse_workbook(period_folder: str, file_path: Path, loaded_at: str):
    master_rows: List[ParsedRow] = []
    qa_rows: List[dict] = []

    wb = load_workbook(filename=file_path, data_only=True)

    for ws in wb.worksheets:
        header_dates = extract_merged_header_values(ws)
        latest_date = header_dates.get("D") or header_dates.get("E") or header_dates.get("B") or header_dates.get("C")

        metric_map = {
            "B": ("absolute", "billion UZS", header_dates.get("B")),
            "C": ("share", "percent", header_dates.get("C")),
            "D": ("absolute", "billion UZS", header_dates.get("D")),
            "E": ("share", "percent", header_dates.get("E")),
            "F": ("growth", "percent", latest_date),
        }

        rows_read = 0
        out_rows = 0
        notes: List[str] = []

        for row_idx in range(5, ws.max_row + 1):
            sector_raw = ws[f"A{row_idx}"].value
            sector = str(sector_raw).strip() if sector_raw is not None else ""
            if not sector:
                continue

            rows_read += 1
            for col in ["B", "C", "D", "E", "F"]:
                metric_type, unit, period_dt = metric_map[col]
                if period_dt is None:
                    notes.append(f"Missing period date for column {col}")
                    continue

                source_cell = f"{col}{row_idx}"
                value = excel_value_to_float(ws[source_cell].value)

                master_rows.append(
                    ParsedRow(
                        report_period_folder=period_folder,
                        period_year=period_dt.year,
                        period_month=period_dt.month,
                        period_date=period_dt.strftime("%Y-%m-%d"),
                        period_label=format_period_label(period_dt),
                        sector=sector,
                        metric_type=metric_type,
                        value=value,
                        unit=unit,
                        source_file=str(file_path.as_posix()),
                        source_sheet=ws.title,
                        source_cell=source_cell,
                        loaded_at=loaded_at,
                    )
                )
                out_rows += 1

        status = "ok" if rows_read > 0 else "warning"
        if rows_read == 0:
            notes.append("No data rows found in column A from row 5 onward")

        qa_rows.append(
            {
                "source_file": str(file_path.as_posix()),
                "report_period_folder": period_folder,
                "source_sheet": ws.title,
                "rows_read": rows_read,
                "output_rows_created": out_rows,
                "status": status,
                "notes": " | ".join(sorted(set(notes))) if notes else "",
            }
        )

    wb.close()
    return master_rows, qa_rows


def write_master(rows: List[ParsedRow]) -> None:
    MASTER_OUT.parent.mkdir(parents=True, exist_ok=True)
    with MASTER_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "sector",
                "metric_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.report_period_folder,
                    row.period_year,
                    row.period_month,
                    row.period_date,
                    row.period_label,
                    row.sector,
                    row.metric_type,
                    "" if row.value is None else row.value,
                    row.unit,
                    row.source_file,
                    row.source_sheet,
                    row.source_cell,
                    row.loaded_at,
                ]
            )


def write_qa(rows: List[dict]) -> None:
    QA_OUT.parent.mkdir(parents=True, exist_ok=True)
    with QA_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def ensure_overwrite_allowed(overwrite: bool) -> None:
    existing = [str(p) for p in [MASTER_OUT, QA_OUT] if p.exists()]
    if existing and not overwrite:
        raise SystemExit(
            "Output file(s) already exist. Use --overwrite to replace: " + ", ".join(existing)
        )


def main() -> None:
    _ = get_column_letter(1)
    args = parse_args()
    ensure_overwrite_allowed(args.overwrite)

    period_folders = build_period_folder_list(args.periods)
    file_list = find_target_files(period_folders)

    loaded_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    all_master: List[ParsedRow] = []
    all_qa: List[dict] = []

    for period_folder, file_path in file_list:
        master_rows, qa_rows = parse_workbook(period_folder, file_path, loaded_at)
        all_master.extend(master_rows)
        all_qa.extend(qa_rows)

    write_master(all_master)
    write_qa(all_qa)

    print(f"Parsed files: {len(file_list)}")
    print(f"Master rows: {len(all_master)} -> {MASTER_OUT}")
    print(f"QA rows: {len(all_qa)} -> {QA_OUT}")


if __name__ == "__main__":
    main()
