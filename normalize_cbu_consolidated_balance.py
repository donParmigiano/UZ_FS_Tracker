"""Parse CBU "Consolidated balance of commercial banks" Excel files into master and QA CSVs.

Run:
    python normalize_cbu_consolidated_balance.py
    python normalize_cbu_consolidated_balance.py --overwrite
    python normalize_cbu_consolidated_balance.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_consolidated_balance_master.csv")
QA_OUTPUT = Path("data/master/cbu_consolidated_balance_parse_qa.csv")
FILE_TOKEN = "consolidated-balance-of-commercial-banks"
PERIOD_FOLDER_RE = re.compile(r"^(\d{4})_(\d{2})$")
SECTION_HEADERS = {"assets", "liabilities", "capital"}


@dataclass
class ParsedRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    balance_section: str
    indicator: str
    metric_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QAEntry:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU consolidated balance Excel reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated folders (YYYY_MM) under data/raw/cbu_bankstats to parse.",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def parse_period_folders_arg(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None
    requested: set[str] = set()
    for raw in periods_arg.split(","):
        token = raw.strip()
        if not PERIOD_FOLDER_RE.match(token):
            raise ValueError(f"Invalid period folder '{token}'. Expected YYYY_MM.")
        requested.add(token)
    return requested


def find_period_folder(file_path: Path) -> Optional[str]:
    for parent in file_path.parents:
        if PERIOD_FOLDER_RE.match(parent.name):
            return parent.name
    return None


def find_input_files(root_dir: Path, requested_periods: Optional[set[str]]) -> list[Path]:
    files: list[Path] = []
    for path in root_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_TOKEN not in path.name.lower():
            continue

        period_folder = find_period_folder(path)
        if period_folder is None:
            continue
        if requested_periods is not None and period_folder not in requested_periods:
            continue

        files.append(path)
    return sorted(files)


def extract_effective_header_value(sheet, row: int, col: int) -> object:
    value = sheet.cell(row=row, column=col).value
    if value is not None:
        return value

    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sheet.cell(row=merged.min_row, column=merged.min_col).value
    return None


def parse_period_date(header_value: object) -> Optional[datetime]:
    if isinstance(header_value, datetime):
        return header_value

    text = clean_text(header_value)
    if not text:
        return None

    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_workbook(file_path: Path, loaded_at: str) -> tuple[list[ParsedRow], QAEntry]:
    report_period_folder = find_period_folder(file_path) or "unknown"

    workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = workbook.active if workbook.active is not None else workbook.worksheets[0]

    output_rows: list[ParsedRow] = []
    rows_read = 0

    date_by_col: dict[int, datetime] = {}
    for col in (2, 3, 4, 5):
        date_cell_value = extract_effective_header_value(sheet, row=3, col=col)
        parsed_dt = parse_period_date(date_cell_value)
        if parsed_dt is not None:
            date_by_col[col] = parsed_dt

    latest_col_dt = date_by_col.get(4) or date_by_col.get(5)

    column_specs = {
        2: ("amount", "billion UZS"),
        3: ("share", "percent"),
        4: ("amount", "billion UZS"),
        5: ("share", "percent"),
    }

    current_section = ""

    for row_idx in range(5, sheet.max_row + 1):
        label = clean_text(sheet.cell(row=row_idx, column=1).value)
        if not label:
            continue

        label_norm = label.lower()
        if label_norm in SECTION_HEADERS:
            current_section = label
            continue

        created_in_row = 0

        for col_idx, (metric_type, unit) in column_specs.items():
            value = parse_number(sheet.cell(row=row_idx, column=col_idx).value)
            if value is None:
                continue
            dt = date_by_col.get(col_idx)
            if dt is None:
                continue

            output_rows.append(
                ParsedRow(
                    report_period_folder=report_period_folder,
                    period_year=dt.year,
                    period_month=dt.month,
                    period_date=f"{dt.year:04d}-{dt.month:02d}-{dt.day:02d}",
                    period_label=f"{dt.strftime('%B')} {dt.day}, {dt.year}",
                    balance_section=current_section,
                    indicator=label,
                    metric_type=metric_type,
                    value=value,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=f"{get_column_letter(col_idx)}{row_idx}",
                    loaded_at=loaded_at,
                )
            )
            created_in_row += 1

        growth_value = parse_number(sheet.cell(row=row_idx, column=6).value)
        if growth_value is not None and latest_col_dt is not None:
            output_rows.append(
                ParsedRow(
                    report_period_folder=report_period_folder,
                    period_year=latest_col_dt.year,
                    period_month=latest_col_dt.month,
                    period_date=f"{latest_col_dt.year:04d}-{latest_col_dt.month:02d}-{latest_col_dt.day:02d}",
                    period_label=f"{latest_col_dt.strftime('%B')} {latest_col_dt.day}, {latest_col_dt.year}",
                    balance_section=current_section,
                    indicator=label,
                    metric_type="growth",
                    value=growth_value,
                    unit="percent",
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=f"{get_column_letter(6)}{row_idx}",
                    loaded_at=loaded_at,
                )
            )
            created_in_row += 1

        if created_in_row > 0:
            rows_read += 1

    status = "ok" if output_rows else "warning"
    notes = "Parsed successfully." if output_rows else "No output rows found."

    qa = QAEntry(
        source_file=str(file_path),
        report_period_folder=report_period_folder,
        source_sheet=sheet.title,
        rows_read=rows_read,
        output_rows_created=len(output_rows),
        status=status,
        notes=notes,
    )

    workbook.close()
    return output_rows, qa


def write_master(rows: list[ParsedRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "balance_section",
                "indicator",
                "metric_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.report_period_folder,
                    row.period_year,
                    row.period_month,
                    row.period_date,
                    row.period_label,
                    row.balance_section,
                    row.indicator,
                    row.metric_type,
                    row.value,
                    row.unit,
                    row.source_file,
                    row.source_sheet,
                    row.source_cell,
                    row.loaded_at,
                ]
            )


def write_qa(rows: list[QAEntry], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.source_file,
                    row.report_period_folder,
                    row.source_sheet,
                    row.rows_read,
                    row.output_rows_created,
                    row.status,
                    row.notes,
                ]
            )


def main() -> None:
    args = parse_args()

    if (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()) and not args.overwrite:
        raise SystemExit(
            "Output file(s) already exist. Use --overwrite to replace:\n"
            f"- {MASTER_OUTPUT}\n"
            f"- {QA_OUTPUT}"
        )

    try:
        requested_periods = parse_period_folders_arg(args.periods)
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    input_files = find_input_files(RAW_ROOT, requested_periods)
    if not input_files:
        raise SystemExit("No matching CBU consolidated balance files found to parse.")

    loaded_at = datetime.now(timezone.utc).isoformat()

    all_rows: list[ParsedRow] = []
    qa_rows: list[QAEntry] = []

    for file_path in input_files:
        rows, qa = parse_workbook(file_path, loaded_at=loaded_at)
        all_rows.extend(rows)
        qa_rows.append(qa)

    write_master(all_rows, MASTER_OUTPUT)
    write_qa(qa_rows, QA_OUTPUT)

    print(f"Parsed files: {len(input_files)}")
    print(f"Master rows: {len(all_rows)} -> {MASTER_OUTPUT}")
    print(f"QA rows: {len(qa_rows)} -> {QA_OUTPUT}")


if __name__ == "__main__":
    main()
