"""Parse CBU 'Liquidity dynamics of banking sector' Excel files.

Run:
    python normalize_cbu_liquidity_dynamics.py
    python normalize_cbu_liquidity_dynamics.py --overwrite
    python normalize_cbu_liquidity_dynamics.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_liquidity_dynamics_master.csv")
QA_OUTPUT = Path("data/master/cbu_liquidity_dynamics_parse_qa.csv")
FILE_NAME_TOKEN = "liquidity-dynamics-of-banking-sector"
PERIOD_DIR_RE = re.compile(r"^(\d{4})_(\d{2})$")


@dataclass
class MasterRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    indicator: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QaRow:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU liquidity dynamics reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they exist.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated YYYY_MM folders to parse, e.g. 2026_04,2024_04",
    )
    return parser.parse_args()


def ensure_outputs_writable(overwrite: bool) -> None:
    existing = [path for path in (MASTER_OUTPUT, QA_OUTPUT) if path.exists()]
    if existing and not overwrite:
        names = ", ".join(str(path) for path in existing)
        raise SystemExit(f"Output file(s) already exist: {names}. Use --overwrite to replace them.")


def parse_periods_arg(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None

    requested: set[str] = set()
    for token in periods_arg.split(","):
        period = token.strip()
        if not period:
            continue
        match = PERIOD_DIR_RE.match(period)
        if not match:
            raise SystemExit(f"Invalid period '{period}'. Expected YYYY_MM.")
        month = int(match.group(2))
        if not 1 <= month <= 12:
            raise SystemExit(f"Invalid period '{period}'. Month must be 01..12.")
        requested.add(period)

    if not requested:
        raise SystemExit("--periods was provided but no valid values were found.")
    return requested


def find_report_period_folder(file_path: Path) -> Optional[str]:
    for parent in file_path.parents:
        if PERIOD_DIR_RE.match(parent.name):
            return parent.name
    return None


def find_input_files(root: Path, requested_periods: Optional[set[str]]) -> list[Path]:
    matches: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_NAME_TOKEN not in path.name.lower():
            continue
        period_folder = find_report_period_folder(path)
        if requested_periods is not None and period_folder not in requested_periods:
            continue
        matches.append(path)
    return sorted(matches)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    compact = text.replace(" ", "")
    if "," in compact and "." in compact:
        compact = compact.replace(",", "")
    elif "," in compact:
        compact = compact.replace(",", ".")

    try:
        return float(compact)
    except ValueError:
        return None


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def format_period_label(dt: date) -> str:
    return f"{dt.strftime('%B')} {dt.day}, {dt.year}"


def normalize_indicator_and_unit(indicator: str) -> tuple[str, str]:
    cleaned = indicator
    lowered = indicator.lower()

    if "(in billion uzs)" in lowered:
        cleaned = re.sub(r"\(in\s+billion\s+uzs\)", "", cleaned, flags=re.IGNORECASE).strip()
        return cleaned, "billion UZS"
    if "(in %)" in lowered:
        cleaned = re.sub(r"\(in\s*%\)", "", cleaned, flags=re.IGNORECASE).strip()
        return cleaned, "percent"

    ratio_terms = (
        "liquidity coverage ratio",
        "net stable funding ratio",
        "immediate liquidity ratio",
    )
    if any(term in lowered for term in ratio_terms):
        return cleaned, "ratio"

    return cleaned, "needs_review"


def parse_file(file_path: Path, loaded_at: str) -> tuple[list[MasterRow], QaRow]:
    period_folder = find_report_period_folder(file_path)
    if period_folder is None:
        return [], QaRow(str(file_path), "unknown", "", 0, 0, "failed", "Could not infer report period folder.")

    folder_match = PERIOD_DIR_RE.match(period_folder)
    if folder_match is None:
        return [], QaRow(str(file_path), period_folder, "", 0, 0, "failed", "Invalid period folder format.")

    period_year = int(folder_match.group(1))
    period_month = int(folder_match.group(2))

    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = wb.active if wb.active is not None else wb.worksheets[0]

    period_dates: dict[str, date] = {}
    for col_idx in (2, 3):
        col_letter = get_column_letter(col_idx)
        dt = parse_excel_date(sheet[f"{col_letter}3"].value)
        if dt is not None:
            period_dates[col_letter] = dt

    if not period_dates:
        wb.close()
        return [], QaRow(str(file_path), period_folder, sheet.title, 0, 0, "failed", "Could not parse period headers from B3/C3.")

    rows_read = 0
    output_rows: list[MasterRow] = []

    for row_idx in range(4, sheet.max_row + 1):
        raw_indicator = clean_text(sheet[f"A{row_idx}"].value)
        if not raw_indicator:
            continue
        if raw_indicator.startswith("*"):
            continue

        rows_read += 1
        indicator, unit = normalize_indicator_and_unit(raw_indicator)

        for col_letter, dt in period_dates.items():
            source_cell = f"{col_letter}{row_idx}"
            value = parse_number(sheet[source_cell].value)
            if value is None:
                continue

            output_rows.append(
                MasterRow(
                    report_period_folder=period_folder,
                    period_year=period_year,
                    period_month=period_month,
                    period_date=dt.isoformat(),
                    period_label=format_period_label(dt),
                    indicator=indicator,
                    value=value,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=source_cell,
                    loaded_at=loaded_at,
                )
            )

    wb.close()
    notes = f"Parsed periods from B3/C3: {', '.join(sorted(period_dates.keys()))}."
    return output_rows, QaRow(str(file_path), period_folder, sheet.title, rows_read, len(output_rows), "ok", notes)


def write_master(rows: list[MasterRow]) -> None:
    MASTER_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with MASTER_OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "indicator",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.report_period_folder,
                    row.period_year,
                    row.period_month,
                    row.period_date,
                    row.period_label,
                    row.indicator,
                    row.value,
                    row.unit,
                    row.source_file,
                    row.source_sheet,
                    row.source_cell,
                    row.loaded_at,
                ]
            )


def write_qa(rows: list[QaRow]) -> None:
    QA_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with QA_OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.source_file,
                    row.report_period_folder,
                    row.source_sheet,
                    row.rows_read,
                    row.output_rows_created,
                    row.status,
                    row.notes,
                ]
            )


def main() -> None:
    args = parse_args()
    ensure_outputs_writable(args.overwrite)
    requested_periods = parse_periods_arg(args.periods)

    files = find_input_files(RAW_ROOT, requested_periods)
    if not files:
        raise SystemExit("No matching files found for token 'Liquidity-dynamics-of-banking-sector'.")

    loaded_at = datetime.now(timezone.utc).isoformat()
    master_rows: list[MasterRow] = []
    qa_rows: list[QaRow] = []

    for file_path in files:
        parsed_rows, qa_row = parse_file(file_path, loaded_at)
        master_rows.extend(parsed_rows)
        qa_rows.append(qa_row)

    write_master(master_rows)
    write_qa(qa_rows)

    print(f"Parsed {len(files)} file(s).")
    print(f"Master rows: {len(master_rows)} -> {MASTER_OUTPUT}")
    print(f"QA rows: {len(qa_rows)} -> {QA_OUTPUT}")


if __name__ == "__main__":
    main()
