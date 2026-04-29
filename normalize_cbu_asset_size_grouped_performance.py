#!/usr/bin/env python3
"""Parse CBU "Main performance indicators ... grouped in terms of asset size" workbooks.

Usage:
  python normalize_cbu_asset_size_grouped_performance.py
  python normalize_cbu_asset_size_grouped_performance.py --overwrite
  python normalize_cbu_asset_size_grouped_performance.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_asset_size_grouped_performance_master.csv")
QA_OUTPUT = Path("data/master/cbu_asset_size_grouped_performance_parse_qa.csv")
TARGET_TOKEN = "Main-performance-indicators-of-commercial-banks_-grouped-in-terms-of-asset-size"
PERIOD_RE = re.compile(r"^\d{4}_\d{2}$")
DATE_RE = re.compile(
    r"as\s+of\s+([A-Za-z]+\s+\d{1,2},\s*\d{4})",
    flags=re.IGNORECASE,
)
RATIO_HINT_RE = re.compile(r"\b(ratio|roa|roe|car)\b|in\s*%|%", flags=re.IGNORECASE)

MASTER_COLUMNS = [
    "report_period_folder",
    "period_year",
    "period_month",
    "period_date",
    "period_label",
    "performance_section",
    "indicator",
    "asset_size_category",
    "metric_type",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "source_cell",
    "loaded_at",
]

QA_COLUMNS = [
    "source_file",
    "report_period_folder",
    "source_sheet",
    "rows_read",
    "output_rows_created",
    "status",
    "notes",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--periods", help="Comma-separated YYYY_MM folders.")
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def parse_periods(raw: str | None) -> set[str] | None:
    if raw is None:
        return None
    periods = {item.strip() for item in raw.split(",") if item.strip()}
    invalid = sorted(p for p in periods if not PERIOD_RE.fullmatch(p))
    if invalid:
        raise SystemExit(f"Invalid --periods values (expected YYYY_MM): {', '.join(invalid)}")
    if not periods:
        raise SystemExit("--periods was provided but no valid values were found.")
    return periods


def discover_files(periods: set[str] | None) -> list[tuple[str, Path]]:
    if not RAW_ROOT.exists():
        return []
    results: list[tuple[str, Path]] = []
    folders = sorted(p for p in RAW_ROOT.iterdir() if p.is_dir() and PERIOD_RE.fullmatch(p.name))
    for folder in folders:
        if periods is not None and folder.name not in periods:
            continue
        for file_path in sorted(folder.glob("*.xlsx")):
            if TARGET_TOKEN in file_path.name:
                results.append((folder.name, file_path))
    return results


def merged_value(ws: Any, row: int, col: int) -> Any:
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mrange in ws.merged_cells.ranges:
        if mrange.min_row <= row <= mrange.max_row and mrange.min_col <= col <= mrange.max_col:
            return ws.cell(row=mrange.min_row, column=mrange.min_col).value
    return None


def detect_report_date(ws: Any) -> datetime:
    for row in range(1, min(15, ws.max_row) + 1):
        for col in range(1, min(6, ws.max_column) + 1):
            text = merged_value(ws, row, col)
            if not isinstance(text, str):
                continue
            match = DATE_RE.search(text)
            if match:
                return datetime.strptime(match.group(1).strip(), "%B %d, %Y")
    raise ValueError("Could not detect reporting date from title text.")


def detect_unit(ws: Any) -> str:
    j2 = merged_value(ws, 2, 10)
    if isinstance(j2, str) and j2.strip():
        return j2.strip()
    for row in range(1, 8):
        for col in range(1, min(12, ws.max_column) + 1):
            value = merged_value(ws, row, col)
            if isinstance(value, str) and "uzs" in value.lower():
                return value.strip()
    return "billion UZS"


def clean_indicator(text: str) -> str:
    s = re.sub(r",?\s*in\s*%\s*$", "", text, flags=re.IGNORECASE).strip()
    return re.sub(r"\s+", " ", s)


def is_section_header(text: str) -> bool:
    normalized = text.strip().lower()
    return normalized in {"assets", "equity and financial performance", "liabilities"}


def is_ratio_indicator(text: str) -> bool:
    return bool(RATIO_HINT_RE.search(text))


def parse_file(period_folder: str, file_path: Path, loaded_at: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    report_dt = detect_report_date(ws)
    period_date = report_dt.date().isoformat()
    period_label = f"{report_dt.strftime('%B')} {report_dt.day}, {report_dt.year}"
    money_unit = detect_unit(ws)

    asset_pairs = [(2, 3), (4, 5), (6, 7), (8, 9), (10, 11)]
    category_by_col: dict[int, str] = {}
    for c1, c2 in asset_pairs:
        label = merged_value(ws, 4, c1)
        if not isinstance(label, str) or not label.strip():
            label = {
                2: "Total",
                4: "up to 3 trillion UZS",
                6: "from 3 to 10 trillion UZS",
                8: "from 10 to 30 trillion UZS",
                10: "30 trillion UZS and above",
            }[c1]
        label = re.sub(r"\s+", " ", str(label).strip())
        category_by_col[c1] = label
        category_by_col[c2] = label

    rows: list[dict[str, Any]] = []
    qa_rows: list[dict[str, Any]] = []
    current_section = ""
    rows_read = 0

    for row_idx in range(1, ws.max_row + 1):
        ind_raw = merged_value(ws, row_idx, 1)
        if not isinstance(ind_raw, str) or not ind_raw.strip():
            continue
        indicator_text = re.sub(r"\s+", " ", ind_raw.strip())
        if is_section_header(indicator_text):
            current_section = indicator_text
            continue
        if not current_section:
            continue

        row_has_data = False
        for c1, c2 in asset_pairs:
            banks_val = ws.cell(row=row_idx, column=c1).value
            metric_val = ws.cell(row=row_idx, column=c2).value
            category = category_by_col.get(c1, "")

            if banks_val is not None:
                row_has_data = True
                rows.append({
                    "report_period_folder": period_folder,
                    "period_year": report_dt.year,
                    "period_month": report_dt.month,
                    "period_date": period_date,
                    "period_label": period_label,
                    "performance_section": current_section,
                    "indicator": clean_indicator(indicator_text),
                    "asset_size_category": category,
                    "metric_type": "number_of_banks",
                    "value": banks_val,
                    "unit": "banks",
                    "source_file": str(file_path),
                    "source_sheet": ws.title,
                    "source_cell": f"{get_column_letter(c1)}{row_idx}",
                    "loaded_at": loaded_at,
                })
            if metric_val is not None:
                row_has_data = True
                is_ratio = is_ratio_indicator(indicator_text)
                rows.append({
                    "report_period_folder": period_folder,
                    "period_year": report_dt.year,
                    "period_month": report_dt.month,
                    "period_date": period_date,
                    "period_label": period_label,
                    "performance_section": current_section,
                    "indicator": clean_indicator(indicator_text),
                    "asset_size_category": category,
                    "metric_type": "ratio" if is_ratio else "absolute",
                    "value": metric_val,
                    "unit": "percent" if is_ratio else money_unit,
                    "source_file": str(file_path),
                    "source_sheet": ws.title,
                    "source_cell": f"{get_column_letter(c2)}{row_idx}",
                    "loaded_at": loaded_at,
                })
        if row_has_data:
            rows_read += 1

    qa_rows.append({
        "source_file": str(file_path),
        "report_period_folder": period_folder,
        "source_sheet": ws.title,
        "rows_read": rows_read,
        "output_rows_created": len(rows),
        "status": "ok" if rows else "warning",
        "notes": "",
    })
    return rows, qa_rows


def write_csv(path: Path, columns: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    if (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()) and not args.overwrite:
        raise SystemExit(
            "Output file(s) already exist. Use --overwrite to replace: "
            f"{MASTER_OUTPUT} and {QA_OUTPUT}"
        )

    periods = parse_periods(args.periods)
    targets = discover_files(periods)
    if not targets:
        raise SystemExit("No matching source files found.")

    loaded_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    all_rows: list[dict[str, Any]] = []
    all_qa: list[dict[str, Any]] = []

    for folder, file_path in targets:
        rows, qa = parse_file(folder, file_path, loaded_at)
        all_rows.extend(rows)
        all_qa.extend(qa)

    write_csv(MASTER_OUTPUT, MASTER_COLUMNS, all_rows)
    write_csv(QA_OUTPUT, QA_COLUMNS, all_qa)

    print(f"Wrote {len(all_rows)} rows -> {MASTER_OUTPUT}")
    print(f"Wrote {len(all_qa)} rows -> {QA_OUTPUT}")


if __name__ == "__main__":
    main()
