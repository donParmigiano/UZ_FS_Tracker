"""Parse CBU 'Information on major indicators of banking sector' Excel files.

Run:
    python normalize_cbu_major_sector_indicators.py
    python normalize_cbu_major_sector_indicators.py --overwrite
    python normalize_cbu_major_sector_indicators.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_major_sector_indicators_master.csv")
QA_OUTPUT = Path("data/master/cbu_major_sector_indicators_parse_qa.csv")
FILE_NAME_TOKEN = "information-on-major-indicators-of-banking-sector"
PERIOD_DIR_RE = re.compile(r"^(\d{4})_(\d{2})$")

METRIC_MAP = {
    "B": "total",
    "C": "foreign_currency",
    "D": "foreign_currency_share",
    "E": "total",
    "F": "foreign_currency",
    "G": "foreign_currency_share",
    "H": "growth",
}

UNIT_MAP = {
    "total": "billion UZS",
    "foreign_currency": "billion UZS",
    "foreign_currency_share": "percent",
    "growth": "percent",
}


@dataclass
class MasterRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    indicator: str
    metric_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QaRow:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU major banking sector indicator reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they exist.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated YYYY_MM folders to parse, e.g. 2026_04,2024_04",
    )
    return parser.parse_args()


def ensure_outputs_writable(overwrite: bool) -> None:
    existing = [path for path in (MASTER_OUTPUT, QA_OUTPUT) if path.exists()]
    if existing and not overwrite:
        names = ", ".join(str(p) for p in existing)
        raise SystemExit(f"Output file(s) already exist: {names}. Use --overwrite to replace them.")


def parse_periods_arg(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None
    requested: set[str] = set()
    for token in periods_arg.split(","):
        period = token.strip()
        if not period:
            continue
        match = PERIOD_DIR_RE.match(period)
        if not match:
            raise SystemExit(f"Invalid period '{period}'. Expected YYYY_MM.")
        month = int(match.group(2))
        if not 1 <= month <= 12:
            raise SystemExit(f"Invalid period '{period}'. Month must be 01..12.")
        requested.add(period)
    if not requested:
        raise SystemExit("--periods was provided but no valid values were found.")
    return requested


def find_report_period_folder(file_path: Path) -> Optional[str]:
    for parent in file_path.parents:
        if PERIOD_DIR_RE.match(parent.name):
            return parent.name
    return None


def find_input_files(root: Path, requested_periods: Optional[set[str]]) -> list[Path]:
    matches: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_NAME_TOKEN not in path.name.lower():
            continue
        folder = find_report_period_folder(path)
        if requested_periods is not None and folder not in requested_periods:
            continue
        matches.append(path)
    return sorted(matches)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    compact = text.replace(" ", "")
    if "," in compact and "." in compact:
        compact = compact.replace(",", "")
    elif "," in compact:
        compact = compact.replace(",", ".")

    try:
        return float(compact)
    except ValueError:
        return None


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def period_label(dt: date) -> str:
    return f"{dt.strftime('%B')} {dt.day}, {dt.year}"


def expand_period_headers(sheet) -> tuple[dict[str, date], Optional[date], str]:
    cell_dates: dict[str, date] = {}
    latest_period: Optional[date] = None

    for merged_range in sheet.merged_cells.ranges:
        if merged_range.min_row == 3 and merged_range.max_row == 3:
            top_left = sheet.cell(row=3, column=merged_range.min_col)
            dt = parse_excel_date(top_left.value)
            if dt is None:
                continue
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                cell_dates[get_column_letter(col_idx)] = dt

    unit_text = clean_text(sheet["H2"].value)

    for col in ("B", "C", "D", "E", "F", "G"):
        dt = cell_dates.get(col)
        if dt is not None and (latest_period is None or dt > latest_period):
            latest_period = dt

    if latest_period is not None:
        cell_dates["H"] = latest_period

    notes = "Parsed merged period headers at row 3."
    if not unit_text:
        notes += " H2 unit text is empty."
    return cell_dates, latest_period, notes


def parse_file(file_path: Path, loaded_at: str) -> tuple[list[MasterRow], QaRow]:
    period_folder = find_report_period_folder(file_path)
    if period_folder is None:
        return [], QaRow(str(file_path), "unknown", "", 0, 0, "failed", "Could not infer report period folder.")

    workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = workbook.active if workbook.active is not None else workbook.worksheets[0]

    col_period_dates, _, notes = expand_period_headers(sheet)
    output_rows: list[MasterRow] = []
    rows_read = 0

    for row_idx in range(4, sheet.max_row + 1):
        indicator = clean_text(sheet[f"A{row_idx}"].value)
        if not indicator:
            continue

        created_for_indicator = 0
        for col_letter, metric_type in METRIC_MAP.items():
            period_dt = col_period_dates.get(col_letter)
            if period_dt is None:
                continue

            source_cell = f"{col_letter}{row_idx}"
            value = parse_number(sheet[source_cell].value)
            if value is None:
                continue

            output_rows.append(
                MasterRow(
                    report_period_folder=period_folder,
                    period_year=period_dt.year,
                    period_month=period_dt.month,
                    period_date=period_dt.isoformat(),
                    period_label=period_label(period_dt),
                    indicator=indicator,
                    metric_type=metric_type,
                    value=value,
                    unit=UNIT_MAP[metric_type],
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=source_cell,
                    loaded_at=loaded_at,
                )
            )
            created_for_indicator += 1

        if created_for_indicator > 0:
            rows_read += 1

    status = "ok" if output_rows else "warning"
    if not output_rows:
        notes = f"{notes} No numeric output rows created."

    qa = QaRow(
        source_file=str(file_path),
        report_period_folder=period_folder,
        source_sheet=sheet.title,
        rows_read=rows_read,
        output_rows_created=len(output_rows),
        status=status,
        notes=notes,
    )
    return output_rows, qa


def write_master(rows: list[MasterRow]) -> None:
    MASTER_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with MASTER_OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "indicator",
                "metric_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.report_period_folder,
                    row.period_year,
                    row.period_month,
                    row.period_date,
                    row.period_label,
                    row.indicator,
                    row.metric_type,
                    row.value,
                    row.unit,
                    row.source_file,
                    row.source_sheet,
                    row.source_cell,
                    row.loaded_at,
                ]
            )


def write_qa(rows: list[QaRow]) -> None:
    QA_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with QA_OUTPUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.source_file,
                    row.report_period_folder,
                    row.source_sheet,
                    row.rows_read,
                    row.output_rows_created,
                    row.status,
                    row.notes,
                ]
            )


def main() -> None:
    args = parse_args()
    ensure_outputs_writable(overwrite=args.overwrite)

    requested_periods = parse_periods_arg(args.periods)
    files = find_input_files(RAW_ROOT, requested_periods)

    if not files:
        requested_msg = "all periods" if requested_periods is None else ", ".join(sorted(requested_periods))
        raise SystemExit(f"No matching files found for {requested_msg}.")

    loaded_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    master_rows: list[MasterRow] = []
    qa_rows: list[QaRow] = []

    for file_path in files:
        rows, qa = parse_file(file_path=file_path, loaded_at=loaded_at)
        master_rows.extend(rows)
        qa_rows.append(qa)

    write_master(master_rows)
    write_qa(qa_rows)

    print(f"Wrote {len(master_rows)} rows to {MASTER_OUTPUT}")
    print(f"Wrote {len(qa_rows)} rows to {QA_OUTPUT}")


if __name__ == "__main__":
    main()
