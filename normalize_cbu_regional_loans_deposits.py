"""Parse recurring CBU report: Information on total loans and total deposits of banking system by regions.

Run:
    python normalize_cbu_regional_loans_deposits.py
    python normalize_cbu_regional_loans_deposits.py --overwrite
    python normalize_cbu_regional_loans_deposits.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_regional_loans_deposits_master.csv")
QA_OUTPUT = Path("data/master/cbu_regional_loans_deposits_parse_qa.csv")
FILE_NAME_CONTAINS = "information-on-total-loans-and-total-deposits-of-banking-system-by-regions"


@dataclass
class MasterRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    region: str
    indicator: str
    customer_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QAEntry:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU regional loans/deposits report into master + QA CSV.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they already exist.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated list of YYYY_MM folders to parse (example: 2026_04,2024_04).",
    )
    return parser.parse_args()


def parse_period_filters(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None
    values = [p.strip() for p in periods_arg.split(",") if p.strip()]
    pattern = re.compile(r"^\d{4}_\d{2}$")
    invalid = [v for v in values if not pattern.match(v)]
    if invalid:
        raise ValueError(f"Invalid --periods entries: {', '.join(invalid)}. Expected YYYY_MM format.")
    return set(values)


def find_period_folder(file_path: Path) -> Optional[str]:
    pattern = re.compile(r"^(\d{4})_(\d{2})$")
    for parent in file_path.parents:
        if pattern.match(parent.name):
            return parent.name
    return None


def find_input_files(root: Path, period_filters: Optional[set[str]]) -> list[Path]:
    files: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_NAME_CONTAINS not in path.name.lower():
            continue
        folder = find_period_folder(path)
        if period_filters is not None and folder not in period_filters:
            continue
        files.append(path)
    return sorted(files)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date_from_title(title: str) -> Optional[datetime]:
    match = re.search(r"as of\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})", title, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%B %d, %Y")
    except ValueError:
        return None


def resolve_cell_value(sheet, row_idx: int, col_idx: int) -> object:
    cell = sheet.cell(row=row_idx, column=col_idx)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row_idx <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
            return sheet.cell(row=merged.min_row, column=merged.min_col).value
    return None


def parse_workbook(file_path: Path, loaded_at: str) -> tuple[list[MasterRow], QAEntry]:
    period_folder = find_period_folder(file_path) or "unknown"
    workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = workbook.active if workbook.active is not None else workbook.worksheets[0]

    title = clean_text(resolve_cell_value(sheet, 1, 1))
    period_dt = parse_date_from_title(title)
    if period_dt is None:
        workbook.close()
        return [], QAEntry(str(file_path), period_folder, sheet.title, 0, 0, "failed", "Could not parse period date from title row.")

    period_year = period_dt.year
    period_month = period_dt.month
    period_date = period_dt.strftime("%Y-%m-%d")
    period_label = f"{period_dt.strftime('%B')} {period_dt.day}, {period_dt.year}"
    unit = clean_text(resolve_cell_value(sheet, 2, 8)) or "billion UZS"

    column_map = {
        3: ("loans", "total"),
        4: ("loans", "individuals"),
        5: ("loans", "legal_entities"),
        6: ("deposits", "total"),
        7: ("deposits", "individuals"),
        8: ("deposits", "legal_entities"),
    }

    rows: list[MasterRow] = []
    rows_read = 0

    for row_idx in range(1, sheet.max_row + 1):
        region = clean_text(resolve_cell_value(sheet, row_idx, 2))
        if not region:
            continue

        numeric_cells: list[tuple[str, str, float, str]] = []
        for col_idx, (indicator, customer_type) in column_map.items():
            value = parse_number(resolve_cell_value(sheet, row_idx, col_idx))
            if value is None:
                continue
            source_cell = f"{get_column_letter(col_idx)}{row_idx}"
            numeric_cells.append((indicator, customer_type, value, source_cell))

        if not numeric_cells:
            continue

        rows_read += 1
        for indicator, customer_type, value, source_cell in numeric_cells:
            rows.append(
                MasterRow(
                    report_period_folder=period_folder,
                    period_year=period_year,
                    period_month=period_month,
                    period_date=period_date,
                    period_label=period_label,
                    region=region,
                    indicator=indicator,
                    customer_type=customer_type,
                    value=value,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=source_cell,
                    loaded_at=loaded_at,
                )
            )

    workbook.close()

    if not rows:
        return [], QAEntry(str(file_path), period_folder, sheet.title, rows_read, 0, "failed", "No numeric rows were parsed.")

    return rows, QAEntry(str(file_path), period_folder, sheet.title, rows_read, len(rows), "ok", "")


def write_master(rows: list[MasterRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "region",
                "indicator",
                "customer_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.report_period_folder,
                    row.period_year,
                    row.period_month,
                    row.period_date,
                    row.period_label,
                    row.region,
                    row.indicator,
                    row.customer_type,
                    row.value,
                    row.unit,
                    row.source_file,
                    row.source_sheet,
                    row.source_cell,
                    row.loaded_at,
                ]
            )


def write_qa(rows: list[QAEntry], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.source_file,
                    row.report_period_folder,
                    row.source_sheet,
                    row.rows_read,
                    row.output_rows_created,
                    row.status,
                    row.notes,
                ]
            )


def main() -> None:
    args = parse_args()

    if (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()) and not args.overwrite:
        raise SystemExit(
            "Output files already exist. Re-run with --overwrite to replace: "
            f"{MASTER_OUTPUT} and {QA_OUTPUT}"
        )

    period_filters = parse_period_filters(args.periods)
    input_files = find_input_files(RAW_ROOT, period_filters)

    if not input_files:
        raise SystemExit("No matching CBU regional loans/deposits files found in data/raw/cbu_bankstats.")

    loaded_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    all_rows: list[MasterRow] = []
    qa_rows: list[QAEntry] = []

    for file_path in input_files:
        parsed_rows, qa = parse_workbook(file_path, loaded_at)
        all_rows.extend(parsed_rows)
        qa_rows.append(qa)

    write_master(all_rows, MASTER_OUTPUT)
    write_qa(qa_rows, QA_OUTPUT)

    print(f"Parsed files: {len(input_files)}")
    print(f"Master rows written: {len(all_rows)} -> {MASTER_OUTPUT}")
    print(f"QA rows written: {len(qa_rows)} -> {QA_OUTPUT}")


if __name__ == "__main__":
    main()
