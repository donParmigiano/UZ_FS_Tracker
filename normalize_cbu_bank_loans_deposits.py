"""Parse recurring CBU report: Information on loans and deposits of banks.

Run:
    python normalize_cbu_bank_loans_deposits.py
    python normalize_cbu_bank_loans_deposits.py --overwrite
    python normalize_cbu_bank_loans_deposits.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_bank_loans_deposits_master.csv")
QA_OUTPUT = Path("data/master/cbu_bank_loans_deposits_parse_qa.csv")
FILE_NAME_CONTAINS = "information-on-loans-and-deposits-of-banks"


@dataclass
class MasterRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    bank_group: str
    bank_name: str
    indicator: str
    customer_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QAEntry:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU loans/deposits Excel files into master + QA CSV.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they already exist.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated list of YYYY_MM folders to parse (example: 2026_04,2024_04).",
    )
    return parser.parse_args()


def parse_period_filters(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None
    values = [p.strip() for p in periods_arg.split(",") if p.strip()]
    pattern = re.compile(r"^\d{4}_\d{2}$")
    invalid = [v for v in values if not pattern.match(v)]
    if invalid:
        raise ValueError(f"Invalid --periods entries: {', '.join(invalid)}. Expected YYYY_MM format.")
    return set(values)


def find_period_folder(file_path: Path) -> Optional[str]:
    pattern = re.compile(r"^(\d{4})_(\d{2})$")
    for parent in file_path.parents:
        if pattern.match(parent.name):
            return parent.name
    return None


def find_input_files(root: Path, period_filters: Optional[set[str]]) -> list[Path]:
    files: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_NAME_CONTAINS not in path.name.lower():
            continue
        folder = find_period_folder(path)
        if period_filters is not None and folder not in period_filters:
            continue
        files.append(path)
    return sorted(files)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_date_from_title(title: str) -> Optional[datetime]:
    match = re.search(r"as of\s+([A-Za-z]+\s+\d{1,2},\s+\d{4})", title, flags=re.IGNORECASE)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%B %d, %Y")
    except ValueError:
        return None


def get_unit_from_h2(sheet) -> str:
    unit_text = clean_text(sheet["H2"].value)
    return unit_text if unit_text else "billion UZS"


def resolve_cell_value(sheet, row_idx: int, col_idx: int) -> object:
    cell = sheet.cell(row=row_idx, column=col_idx)
    if not isinstance(cell, MergedCell):
        return cell.value
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row_idx <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
            return sheet.cell(row=merged.min_row, column=merged.min_col).value
    return None


def parse_workbook(file_path: Path, loaded_at: str) -> tuple[list[MasterRow], QAEntry]:
    period_folder = find_period_folder(file_path) or "unknown"
    workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = workbook.active if workbook.active is not None else workbook.worksheets[0]

    title = clean_text(resolve_cell_value(sheet, 1, 1))
    period_dt = parse_date_from_title(title)
    if period_dt is None:
        qa = QAEntry(str(file_path), period_folder, sheet.title, 0, 0, "failed", "Could not parse period date from title row.")
        workbook.close()
        return [], qa

    period_year = period_dt.year
    period_month = period_dt.month
    period_date = period_dt.strftime("%Y-%m-%d")
    period_label = f"{period_dt.strftime('%B')} {period_dt.day}, {period_dt.year}"
    unit = get_unit_from_h2(sheet)

    column_map = {
        3: ("loans", "total"),
        4: ("loans", "individuals"),
        5: ("loans", "legal_entities"),
        6: ("deposits", "total"),
        7: ("deposits", "individuals"),
        8: ("deposits", "legal_entities"),
    }

    rows: list[MasterRow] = []
    rows_read = 0
    current_group = ""

    for row_idx in range(1, sheet.max_row + 1):
        col_a = clean_text(resolve_cell_value(sheet, row_idx, 1))
        col_b = clean_text(resolve_cell_value(sheet, row_idx, 2))

        lower_a = col_a.lower()
        is_total = lower_a == "total"
        is_state_group = lower_a == "banks with state ownership"
        is_other_group = lower_a == "other banks"

        if is_state_group:
            current_group = "Banks with State ownership"
        elif is_other_group:
            current_group = "Other banks"
        elif is_total:
            current_group = "Total"

        numeric_cells: list[tuple[str, str, float, str]] = []
        for col_idx, (indicator, customer_type) in column_map.items():
            value = parse_number(resolve_cell_value(sheet, row_idx, col_idx))
            if value is None:
                continue
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            numeric_cells.append((indicator, customer_type, value, cell_ref))

        if not numeric_cells:
            continue

        bank_group = current_group
        bank_name = ""

        if is_total:
            bank_group = "Total"
            bank_name = "Total"
        elif is_state_group:
            bank_group = "Banks with State ownership"
            bank_name = "Banks with State ownership"
        elif is_other_group:
            bank_group = "Other banks"
            bank_name = "Other banks"
        elif col_b:
            bank_name = col_b
            if not bank_group:
                bank_group = "Unspecified"
        else:
            continue

        rows_read += 1
        for indicator, customer_type, value, cell_ref in numeric_cells:
            rows.append(
                MasterRow(
                    report_period_folder=period_folder,
                    period_year=period_year,
                    period_month=period_month,
                    period_date=period_date,
                    period_label=period_label,
                    bank_group=bank_group,
                    bank_name=bank_name,
                    indicator=indicator,
                    customer_type=customer_type,
                    value=value,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=cell_ref,
                    loaded_at=loaded_at,
                )
            )

    workbook.close()
    status = "ok" if rows else "warning"
    notes = "Parsed successfully." if rows else "No numeric rows in C:H for recognized entities."
    return rows, QAEntry(str(file_path), period_folder, sheet.title, rows_read, len(rows), status, notes)


def write_master(rows: list[MasterRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "bank_group",
                "bank_name",
                "indicator",
                "customer_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ]
        )
        for row in rows:
            writer.writerow([
                row.report_period_folder,
                row.period_year,
                row.period_month,
                row.period_date,
                row.period_label,
                row.bank_group,
                row.bank_name,
                row.indicator,
                row.customer_type,
                row.value,
                row.unit,
                row.source_file,
                row.source_sheet,
                row.source_cell,
                row.loaded_at,
            ])


def write_qa(entries: list[QAEntry], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "source_file",
            "report_period_folder",
            "source_sheet",
            "rows_read",
            "output_rows_created",
            "status",
            "notes",
        ])
        for e in entries:
            writer.writerow([e.source_file, e.report_period_folder, e.source_sheet, e.rows_read, e.output_rows_created, e.status, e.notes])


def main() -> int:
    args = parse_args()

    if (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()) and not args.overwrite:
        print(
            "Output file(s) already exist. Use --overwrite to replace:\n"
            f"- {MASTER_OUTPUT}\n"
            f"- {QA_OUTPUT}"
        )
        return 1

    try:
        periods = parse_period_filters(args.periods)
    except ValueError as exc:
        print(str(exc))
        return 1

    files = find_input_files(RAW_ROOT, periods)
    if not files:
        print("No matching files found for Information-on-loans-and-deposits-of-banks.")
        return 1

    loaded_at = datetime.now(timezone.utc).isoformat()
    all_rows: list[MasterRow] = []
    qa_entries: list[QAEntry] = []

    for file_path in files:
        parsed_rows, qa = parse_workbook(file_path, loaded_at)
        all_rows.extend(parsed_rows)
        qa_entries.append(qa)

    write_master(all_rows, MASTER_OUTPUT)
    write_qa(qa_entries, QA_OUTPUT)

    print(f"Wrote {len(all_rows)} rows to {MASTER_OUTPUT}")
    print(f"Wrote {len(qa_entries)} QA rows to {QA_OUTPUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
