#!/usr/bin/env python3
"""Normalize CBU deposits by maturity report files into master and QA CSV outputs."""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUT = Path("data/master/cbu_deposits_by_maturity_master.csv")
QA_OUT = Path("data/master/cbu_deposits_by_maturity_parse_qa.csv")
TARGET_TOKEN = "The-amount-of-deposits-by-maturities"

MASTER_COLUMNS = [
    "report_period_folder",
    "period_year",
    "period_month",
    "period_date",
    "period_label",
    "maturity_bucket",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "source_cell",
    "loaded_at",
]

QA_COLUMNS = [
    "source_file",
    "report_period_folder",
    "source_sheet",
    "rows_read",
    "output_rows_created",
    "status",
    "notes",
]

PERIOD_RE = re.compile(r"^\d{4}_\d{2}$")


@dataclass
class PeriodInfo:
    year: int
    month: int
    dt: date


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU deposits by maturity report files.")
    parser.add_argument("--periods", help="Comma-separated YYYY_MM folders under data/raw/cbu_bankstats")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files")
    return parser.parse_args()


def ensure_outputs_writable(overwrite: bool) -> None:
    if overwrite:
        return
    existing = [str(p) for p in (MASTER_OUT, QA_OUT) if p.exists()]
    if existing:
        raise SystemExit(
            "Output file(s) already exist. Use --overwrite to replace: " + ", ".join(existing)
        )


def parse_folder_period(folder: str) -> Optional[PeriodInfo]:
    if not PERIOD_RE.match(folder):
        return None
    year_s, month_s = folder.split("_")
    try:
        dt = date(int(year_s), int(month_s), 1)
    except ValueError:
        return None
    return PeriodInfo(year=dt.year, month=dt.month, dt=dt)


def list_target_files(periods_arg: Optional[str]) -> List[Tuple[Path, str]]:
    if periods_arg:
        requested = [p.strip() for p in periods_arg.split(",") if p.strip()]
        bad = [p for p in requested if not PERIOD_RE.match(p)]
        if bad:
            raise SystemExit(f"Invalid --periods values (expected YYYY_MM): {', '.join(bad)}")
        folders = [RAW_ROOT / p for p in requested]
    else:
        if not RAW_ROOT.exists():
            return []
        folders = [p for p in RAW_ROOT.iterdir() if p.is_dir()]

    targets: List[Tuple[Path, str]] = []
    for folder in sorted(folders):
        if not folder.exists():
            continue
        folder_name = folder.name
        for file_path in sorted(folder.iterdir()):
            if not file_path.is_file():
                continue
            if TARGET_TOKEN not in file_path.name:
                continue
            if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                continue
            targets.append((file_path, folder_name))
    return targets


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def header_map(sheet) -> dict:
    mapping = {}
    for col_idx in range(2, 8):
        col = get_column_letter(col_idx)
        raw = sheet.cell(row=4, column=col_idx).value
        label = clean_text(raw)
        if not label:
            continue
        mapping[col_idx] = label
    return mapping


def parse_workbook(file_path: Path, report_period_folder: str, loaded_at: str) -> Tuple[List[dict], List[dict]]:
    wb = load_workbook(file_path, data_only=True)
    master_rows: List[dict] = []
    qa_rows: List[dict] = []

    for sheet in wb.worksheets:
        rows_read = 0
        output_rows = 0
        notes: List[str] = []
        status = "ok"

        unit = clean_text(sheet["G2"].value) or "billion UZS"
        maturity_by_col = header_map(sheet)
        if not maturity_by_col:
            notes.append("No maturity headers found in row 4 (B:G)")

        max_row = sheet.max_row
        for row_idx in range(5, max_row + 1):
            period_dt = parse_excel_date(sheet.cell(row=row_idx, column=1).value)
            if period_dt is None:
                continue
            rows_read += 1

            for col_idx in range(2, 8):
                bucket = maturity_by_col.get(col_idx)
                if not bucket:
                    continue
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is None or clean_text(value) == "":
                    continue
                col = get_column_letter(col_idx)
                master_rows.append(
                    {
                        "report_period_folder": report_period_folder,
                        "period_year": period_dt.year,
                        "period_month": period_dt.month,
                        "period_date": period_dt.isoformat(),
                        "period_label": f"{period_dt.strftime('%B')} {period_dt.day}, {period_dt.year}",
                        "maturity_bucket": bucket,
                        "value": value,
                        "unit": unit,
                        "source_file": file_path.name,
                        "source_sheet": sheet.title,
                        "source_cell": f"{col}{row_idx}",
                        "loaded_at": loaded_at,
                    }
                )
                output_rows += 1

        if rows_read == 0:
            status = "warning"
            notes.append("No dated rows found in column A")

        qa_rows.append(
            {
                "source_file": file_path.name,
                "report_period_folder": report_period_folder,
                "source_sheet": sheet.title,
                "rows_read": rows_read,
                "output_rows_created": output_rows,
                "status": status,
                "notes": "; ".join(notes),
            }
        )

    wb.close()
    return master_rows, qa_rows


def write_csv(path: Path, columns: List[str], rows: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    ensure_outputs_writable(args.overwrite)

    targets = list_target_files(args.periods)
    if not targets:
        raise SystemExit("No matching CBU deposits by maturity files found to parse.")

    loaded_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_master: List[dict] = []
    all_qa: List[dict] = []

    for file_path, period_folder in targets:
        master_rows, qa_rows = parse_workbook(file_path, period_folder, loaded_at)
        all_master.extend(master_rows)
        all_qa.extend(qa_rows)

    write_csv(MASTER_OUT, MASTER_COLUMNS, all_master)
    write_csv(QA_OUT, QA_COLUMNS, all_qa)

    print(f"Parsed files: {len(targets)}")
    print(f"Master rows: {len(all_master)} -> {MASTER_OUT}")
    print(f"QA rows: {len(all_qa)} -> {QA_OUT}")


if __name__ == "__main__":
    main()
