"""Collect historical CBU major bank indicators Excel files.

Run:
    python collect_cbu_major_bank_indicators_history.py
    python collect_cbu_major_bank_indicators_history.py --year-start 2019 --year-end 2026 --workers 5
    python collect_cbu_major_bank_indicators_history.py --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from urllib.parse import urlencode, urljoin, urlparse

import requests
import pandas as pd
from bs4 import BeautifulSoup
from bs4.element import Tag

BASE_URL = "https://cbu.uz"
BANKSTATS_PATH = "/en/statistics/bankstats/"
SECTION_ID = "3497"
DEFAULT_FILTER_YEAR = 2026
DEFAULT_FILTER_MONTH = 1
RAW_ROOT = Path("data/raw/cbu_bankstats")
REPORT_PATH = Path("data/processed/cbu_major_bank_indicators_collection_report.csv")
SUMMARY_PATH = Path("data/processed/cbu_major_bank_indicators_collection_summary.csv")


@dataclass
class CollectionRow:
    period_year: int
    period_month: int
    period: str
    listing_url: str
    report_page_url: str
    report_title_detected: str
    excel_file_url: str
    source_method: str
    html_fallback_status: str
    html_fallback_path: str
    local_file_path: str
    status: str
    error_message: str
    collected_at: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect historical CBU major bank indicator files.")
    parser.add_argument("--year-start", type=int, default=2018)
    parser.add_argument("--year-end", type=int, default=datetime.now(timezone.utc).year)
    parser.add_argument("--workers", type=int, default=5)
    parser.add_argument("--overwrite", action="store_true")
    return parser.parse_args()


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "UZ-FS-Tracker/1.0 (+historical-collector; respectful-crawl)",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
    )
    return session


def fetch_text(session: requests.Session, url: str, timeout: float = 12.0, retries: int = 3) -> str:
    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
            return response.text
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(0.5 * (attempt + 1))
    raise RuntimeError(f"Failed to fetch URL: {url}; error={last_error}")


def fetch_bytes(session: requests.Session, url: str, timeout: float = 20.0, retries: int = 3) -> bytes:
    last_error: Optional[Exception] = None
    for attempt in range(retries):
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
            return response.content
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            time.sleep(0.7 * (attempt + 1))
    if last_error is not None:
        raise last_error
    raise RuntimeError(f"Failed to download URL: {url}; error=unknown")


def slugify(text: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return cleaned[:120] if cleaned else "report"


def extract_report_id(page_url: str) -> str:
    match = re.search(r"/bankstats/(\d+)/", urlparse(page_url).path)
    return match.group(1) if match else "unknown"


def select_largest_html_table(page_html: str) -> Optional[pd.DataFrame]:
    try:
        tables = pd.read_html(page_html)
    except Exception:  # noqa: BLE001
        tables = []
    valid_tables = [df for df in tables if not df.empty and df.shape[0] > 0 and df.shape[1] > 0]
    if valid_tables:
        return max(valid_tables, key=lambda df: df.shape[0] * df.shape[1])

    soup = BeautifulSoup(page_html, "html.parser")
    best_df: Optional[pd.DataFrame] = None
    best_score = 0
    for table in soup.find_all("table"):
        rows_data: list[list[str]] = []
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            row = [cell.get_text(" ", strip=True) for cell in cells]
            if any(cell for cell in row):
                rows_data.append(row)
        if not rows_data:
            continue
        max_cols = max(len(r) for r in rows_data)
        if max_cols == 0:
            continue
        normalized_rows = [r + [""] * (max_cols - len(r)) for r in rows_data]
        candidate_df = pd.DataFrame(normalized_rows)
        if candidate_df.empty:
            continue
        score = candidate_df.shape[0] * candidate_df.shape[1]
        if score > best_score:
            best_score = score
            best_df = candidate_df
    return best_df



def parse_html_table_grid(table_tag: Tag) -> tuple[list[list[dict[str, object]]], int, int]:
    grid: list[list[dict[str, object]]] = []
    occupied: dict[tuple[int, int], bool] = {}
    max_col = 0

    rows = table_tag.find_all("tr")
    for row_idx, tr in enumerate(rows):
        col_idx = 0
        while occupied.get((row_idx, col_idx), False):
            col_idx += 1

        cells = tr.find_all(["th", "td"], recursive=False)
        for cell in cells:
            while occupied.get((row_idx, col_idx), False):
                col_idx += 1

            rowspan = int(cell.get("rowspan", 1) or 1)
            colspan = int(cell.get("colspan", 1) or 1)
            cell_obj = {
                "row": row_idx + 1,
                "col": col_idx + 1,
                "rowspan": rowspan,
                "colspan": colspan,
                "text": cell.get_text(" ", strip=True),
                "is_header": cell.name == "th",
            }

            while len(grid) <= row_idx:
                grid.append([])
            grid[row_idx].append(cell_obj)

            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    occupied[(r, c)] = True

            col_idx += colspan
            max_col = max(max_col, col_idx)

    row_count = len(rows)
    return grid, row_count, max_col


def select_best_table_tag(page_html: str) -> tuple[Optional[BeautifulSoup], int, int, int]:
    soup = BeautifulSoup(page_html, "html.parser")
    best_table = None
    best_rows = 0
    best_cols = 0
    best_score = 0

    for table in soup.find_all("table"):
        grid, rows, cols = parse_html_table_grid(table)
        if rows == 0 or cols == 0:
            continue
        non_empty_cells = sum(1 for row in grid for cell in row if str(cell.get("text", "")).strip())
        if non_empty_cells == 0:
            continue
        score = rows * cols
        if score > best_score:
            best_table = table
            best_rows = rows
            best_cols = cols
            best_score = score

    return best_table, best_rows, best_cols, best_score


def create_preserved_fallback_workbook(
    fallback_path: Path,
    table_tag: BeautifulSoup,
    flat_df: pd.DataFrame,
    page_url: str,
    page_title: str,
    failed_excel_url: str,
    selected_rows: int,
    selected_cols: int,
    selected_score: int,
) -> None:
    wb = Workbook()
    ws_preserved = wb.active
    ws_preserved.title = "raw_table_preserved"
    ws_flat = wb.create_sheet("raw_table_flat")
    ws_meta = wb.create_sheet("metadata")

    table_grid, _, _ = parse_html_table_grid(table_tag)

    for row in table_grid:
        for cell in row:
            row_num = int(cell["row"])
            col_num = int(cell["col"])
            rowspan = int(cell["rowspan"])
            colspan = int(cell["colspan"])
            text = str(cell["text"])
            is_header = bool(cell["is_header"])

            excel_cell = ws_preserved.cell(row=row_num, column=col_num, value=text)
            excel_cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="center" if (is_header or rowspan > 1 or colspan > 1) else "left",
            )
            if is_header:
                excel_cell.font = Font(bold=True)

            if rowspan > 1 or colspan > 1:
                ws_preserved.merge_cells(
                    start_row=row_num,
                    start_column=col_num,
                    end_row=row_num + rowspan - 1,
                    end_column=col_num + colspan - 1,
                )

    if selected_cols > 0:
        for col_idx in range(1, selected_cols + 1):
            ws_preserved.column_dimensions[get_column_letter(col_idx)].width = 18

    for col_idx, col_name in enumerate(flat_df.columns, start=1):
        ws_flat.cell(row=1, column=col_idx, value=str(col_name)).font = Font(bold=True)
    for row_idx, row_values in enumerate(flat_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_values, start=1):
            ws_flat.cell(row=row_idx, column=col_idx, value="" if pd.isna(value) else str(value))

    metadata_rows = [
        ("report_page_url", page_url),
        ("report_title_detected", page_title),
        ("failed_excel_url", failed_excel_url),
        ("fallback_created_at", datetime.now(timezone.utc).isoformat()),
        ("source_method", "html_table_fallback"),
        ("selected_table_rows", selected_rows),
        ("selected_table_columns", selected_cols),
        ("selected_table_score", selected_score),
    ]
    for idx, (k, v) in enumerate(metadata_rows, start=1):
        ws_meta.cell(row=idx, column=1, value=k)
        ws_meta.cell(row=idx, column=2, value=v)

    wb.save(fallback_path)


def build_listing_url(year: int, month: int) -> str:
    month_str = f"{month:02d}"
    next_year = year + 1 if month == 12 else year
    next_month = 1 if month == 12 else month + 1
    query_items = [
        ("year", f"{DEFAULT_FILTER_YEAR:04d}"),
        ("month", f"{DEFAULT_FILTER_MONTH:02d}"),
        ("arFilter_DATE_ACTIVE_FROM_1", f"01.{month_str}.{year:04d}"),
        ("arFilter_DATE_ACTIVE_FROM_2", f"01.{next_month:02d}.{next_year:04d}"),
        ("arFilter_ff[SECTION_ID]", SECTION_ID),
        ("year", f"{year:04d}"),
        ("month", month_str),
        ("set_filter", ""),
        ("set_filter", "Y"),
    ]
    return f"{BASE_URL}{BANKSTATS_PATH}?{urlencode(query_items)}"


def normalize_report_page_url(url: str, base_url: str) -> Optional[str]:
    joined_url = urljoin(base_url, url)
    parsed = urlparse(joined_url)
    normalized_path = re.sub(r"/+", "/", parsed.path)
    match = re.search(r"/(?:en/)?statistics/bankstats/(\d+)/?", normalized_path)
    if not match:
        return None
    report_id = match.group(1)
    language_prefix = "en/" if normalized_path.startswith("/en/") else ""
    canonical_path = f"/{language_prefix}statistics/bankstats/{report_id}/"
    return parsed._replace(path=canonical_path, params="", query="", fragment="").geturl()


def is_valid_report_page(url: str) -> bool:
    path = urlparse(url).path
    return bool(re.fullmatch(r"/(?:en/)?statistics/bankstats/\d+/", path))


def extract_candidate_report_links(listing_html: str, listing_url: str) -> list[str]:
    soup = BeautifulSoup(listing_html, "html.parser")
    raw_links: list[tuple[str, str]] = []

    for a_tag in soup.find_all("a", href=True):
        href = a_tag.get("href", "")
        raw_links.append((href, a_tag.get_text(" ", strip=True)))

    for match in re.findall(r"""(?:"|')(\/(?:en\/)?statistics\/bankstats\/\d+\/?(?:\?[^"'#]*)?(?:#[^"']*)?)(?:"|')""", listing_html):
        raw_links.append((match, ""))

    unique_urls: dict[str, str] = {}
    duplicate_urls_removed = 0
    raw_report_links_found = 0

    for raw_href, title_text in raw_links:
        normalized_url = normalize_report_page_url(raw_href, listing_url)
        if not normalized_url:
            continue
        if not is_valid_report_page(urlparse(normalized_url)._replace(query="", fragment="").geturl()):
            continue
        raw_report_links_found += 1
        if normalized_url in unique_urls:
            duplicate_urls_removed += 1
            continue
        unique_urls[normalized_url] = title_text

    found = sorted(unique_urls.keys())
    print(f"[SCAN] listing={listing_url}")
    print(f"[SCAN] raw_report_links_found={raw_report_links_found}")
    print(f"[SCAN] unique_report_urls_kept={len(found)}")
    print(f"[SCAN] duplicate_report_urls_removed={duplicate_urls_removed}")
    print(f"[SCAN] report_page_urls={found}")
    return found


def extract_page_title(soup: BeautifulSoup) -> str:
    if soup.title and soup.title.get_text(strip=True):
        return soup.title.get_text(strip=True)
    h1 = soup.find("h1")
    if h1:
        return h1.get_text(" ", strip=True)
    return ""


def extract_excel_links(page_html: str, page_url: str) -> tuple[str, list[str]]:
    soup = BeautifulSoup(page_html, "html.parser")
    page_title = extract_page_title(soup)
    excel_links: set[str] = set()
    for a_tag in soup.find_all("a", href=True):
        href = a_tag.get("href", "")
        full = urljoin(page_url, href)
        path = urlparse(full).path.lower()
        if path.endswith(".xlsx") or path.endswith(".xls"):
            excel_links.add(full)
    if not excel_links:
        for match in re.findall(r"""(?:"|')(https?://[^\s"'<>]+?\.(?:xlsx|xls)(?:\?[^\s"'<>]*)?|/(?:[^\s"'<>]+?\.xls(?:x)?(?:\?[^\s"'<>]*)?))(?:["'])""", page_html, flags=re.IGNORECASE):
            excel_links.add(urljoin(page_url, match))

    return page_title, sorted(excel_links)


def filename_from_url(url: str, page_url: str) -> str:
    report_id = extract_report_id(page_url)
    path = urlparse(url).path
    raw_name = Path(path).name

    if raw_name:
        parsed_name = Path(raw_name)
        stem = parsed_name.stem.strip()
        extension = parsed_name.suffix.strip()

        stem_is_safe = bool(re.fullmatch(r"[A-Za-z0-9._-]+", stem))
        extension_is_safe = extension.lower() in {".xlsx", ".xls"}
        if stem and stem_is_safe and extension_is_safe:
            return f"{stem}_{report_id}{extension.lower()}"

    return f"report_{report_id}.xlsx"


def try_html_fallback(
    *,
    page_html: str,
    out_dir: Path,
    page_url: str,
    page_title: str,
    excel_file_url: str,
    overwrite: bool,
    no_excel_message_prefix: str = "",
    download_exc: Optional[Exception] = None,
) -> CollectionRow:
    fallback_df = select_largest_html_table(page_html)
    selected_table_tag, selected_rows, selected_cols, selected_score = select_best_table_tag(page_html)
    report_id = extract_report_id(page_url)
    fallback_path = out_dir / f"{slugify(page_title)}_{report_id}_html_fallback.xlsx"

    status_prefix = no_excel_message_prefix.strip()
    if status_prefix:
        status_prefix = f"{status_prefix} "

    if fallback_df is None or selected_table_tag is None:
        error_message = "No usable HTML table found."
        if download_exc is not None:
            error_message = f"Excel download failed ({download_exc}); no HTML table available."
        elif status_prefix:
            error_message = "No Excel link found and no usable HTML table found."
        return CollectionRow(
            period_year=0,
            period_month=0,
            period="",
            listing_url="",
            report_page_url=page_url,
            report_title_detected=page_title,
            excel_file_url=excel_file_url,
            source_method="html_table_fallback",
            html_fallback_status="no_html_table_found",
            html_fallback_path="",
            local_file_path="",
            status="no_html_table_found",
            error_message=error_message,
            collected_at="",
        )

    if fallback_path.exists() and not overwrite:
        error_message = "Reused existing HTML fallback file."
        if download_exc is not None:
            error_message = f"Excel download failed ({download_exc}); reused existing HTML fallback file."
        elif status_prefix:
            error_message = "No Excel link found; reused existing HTML fallback file."
        return CollectionRow(
            period_year=0,
            period_month=0,
            period="",
            listing_url="",
            report_page_url=page_url,
            report_title_detected=page_title,
            excel_file_url=excel_file_url,
            source_method="html_table_fallback",
            html_fallback_status="existing",
            html_fallback_path=str(fallback_path),
            local_file_path=str(fallback_path),
            status="html_fallback_skipped_existing",
            error_message=error_message,
            collected_at="",
        )

    create_preserved_fallback_workbook(
        fallback_path=fallback_path,
        table_tag=selected_table_tag,
        flat_df=fallback_df,
        page_url=page_url,
        page_title=page_title,
        failed_excel_url=excel_file_url,
        selected_rows=selected_rows,
        selected_cols=selected_cols,
        selected_score=selected_score,
    )

    error_message = "Created fallback from HTML table."
    if download_exc is not None:
        status_code = None
        if isinstance(download_exc, requests.HTTPError) and download_exc.response is not None:
            status_code = download_exc.response.status_code
        err_suffix = f" (excel_http_status={status_code})" if status_code == 404 else ""
        error_message = f"Excel download failed ({download_exc}){err_suffix}; created fallback from HTML table."
    elif status_prefix:
        error_message = "No Excel link found; created fallback from HTML table."

    return CollectionRow(
        period_year=0,
        period_month=0,
        period="",
        listing_url="",
        report_page_url=page_url,
        report_title_detected=page_title,
        excel_file_url=excel_file_url,
        source_method="html_table_fallback",
        html_fallback_status="created",
        html_fallback_path=str(fallback_path),
        local_file_path=str(fallback_path),
        status="html_fallback_created",
        error_message=error_message,
        collected_at="",
    )


def collect_period(year: int, month: int, overwrite: bool) -> list[CollectionRow]:
    session = create_session()
    now = datetime.now(timezone.utc).isoformat()
    period = f"{year:04d}-{month:02d}"
    listing_url = build_listing_url(year, month)

    try:
        listing_html = fetch_text(session, listing_url)
    except Exception as exc:  # noqa: BLE001
        return [
            CollectionRow(
                period_year=year,
                period_month=month,
                period=period,
                listing_url=listing_url,
                report_page_url="",
                report_title_detected="",
                excel_file_url="",
                source_method="",
                html_fallback_status="",
                html_fallback_path="",
                local_file_path="",
                status="error",
                error_message=str(exc),
                collected_at=now,
            )
        ]

    candidates = extract_candidate_report_links(listing_html, listing_url)
    rows: list[CollectionRow] = []

    for page_url in candidates:
        try:
            page_html = fetch_text(session, page_url)
            page_title, excel_links = extract_excel_links(page_html, page_url)
            print(f"[PAGE] url={page_url} excel_links_found={len(excel_links)}")
            if not excel_links:
                out_dir = RAW_ROOT / f"{year:04d}_{month:02d}"
                out_dir.mkdir(parents=True, exist_ok=True)
                fallback_row = try_html_fallback(
                    page_html=page_html,
                    out_dir=out_dir,
                    page_url=page_url,
                    page_title=page_title,
                    excel_file_url="",
                    overwrite=overwrite,
                    no_excel_message_prefix="No Excel link found",
                )
                rows.append(
                    CollectionRow(
                        period_year=year,
                        period_month=month,
                        period=period,
                        listing_url=listing_url,
                        report_page_url=page_url,
                        report_title_detected=page_title,
                        excel_file_url="",
                        source_method=fallback_row.source_method,
                        html_fallback_status=fallback_row.html_fallback_status,
                        html_fallback_path=fallback_row.html_fallback_path,
                        local_file_path=fallback_row.local_file_path,
                        status=fallback_row.status,
                        error_message=fallback_row.error_message,
                        collected_at=now,
                    )
                )
                continue

            for file_url in excel_links:
                out_dir = RAW_ROOT / f"{year:04d}_{month:02d}"
                out_dir.mkdir(parents=True, exist_ok=True)
                local_path = out_dir / filename_from_url(file_url, page_url)

                if local_path.exists() and not overwrite:
                    rows.append(
                        CollectionRow(
                            period_year=year,
                            period_month=month,
                            period=period,
                            listing_url=listing_url,
                            report_page_url=page_url,
                            report_title_detected=page_title,
                            excel_file_url=file_url,
                            source_method="excel_download",
                            html_fallback_status="",
                            html_fallback_path="",
                            local_file_path=str(local_path),
                            status="skipped_existing",
                            error_message="",
                            collected_at=now,
                        )
                    )
                    continue

                try:
                    blob = fetch_bytes(session, file_url)
                    local_path.write_bytes(blob)
                    rows.append(
                        CollectionRow(
                            period_year=year,
                            period_month=month,
                            period=period,
                            listing_url=listing_url,
                            report_page_url=page_url,
                            report_title_detected=page_title,
                            excel_file_url=file_url,
                            source_method="excel_download",
                            html_fallback_status="",
                            html_fallback_path="",
                            local_file_path=str(local_path),
                            status="downloaded",
                            error_message="",
                            collected_at=now,
                        )
                    )
                    time.sleep(0.2)
                except Exception as download_exc:  # noqa: BLE001
                    try:
                        fallback_row = try_html_fallback(
                            page_html=page_html,
                            out_dir=out_dir,
                            page_url=page_url,
                            page_title=page_title,
                            excel_file_url=file_url,
                            overwrite=overwrite,
                            download_exc=download_exc,
                        )
                        rows.append(
                            CollectionRow(
                                period_year=year,
                                period_month=month,
                                period=period,
                                listing_url=listing_url,
                                report_page_url=page_url,
                                report_title_detected=page_title,
                                excel_file_url=file_url,
                                source_method=fallback_row.source_method,
                                html_fallback_status=fallback_row.html_fallback_status,
                                html_fallback_path=fallback_row.html_fallback_path,
                                local_file_path=fallback_row.local_file_path,
                                status=fallback_row.status,
                                error_message=fallback_row.error_message,
                                collected_at=now,
                            )
                        )
                    except Exception as fallback_exc:  # noqa: BLE001
                        rows.append(
                            CollectionRow(
                                period_year=year,
                                period_month=month,
                                period=period,
                                listing_url=listing_url,
                                report_page_url=page_url,
                                report_title_detected=page_title,
                                excel_file_url=file_url,
                                source_method="html_table_fallback",
                                html_fallback_status="error",
                                html_fallback_path="",
                                local_file_path="",
                                status="error",
                                error_message=f"Excel download failed ({download_exc}); HTML fallback failed ({fallback_exc})",
                                collected_at=now,
                            )
                        )

        except Exception as page_exc:  # noqa: BLE001
            rows.append(
                CollectionRow(
                    period_year=year,
                    period_month=month,
                    period=period,
                    listing_url=listing_url,
                    report_page_url=page_url,
                    report_title_detected="",
                    excel_file_url="",
                    source_method="",
                    html_fallback_status="",
                    html_fallback_path="",
                    local_file_path="",
                    status="error",
                    error_message=str(page_exc),
                    collected_at=now,
                )
            )

    if not rows:
        rows.append(
            CollectionRow(
                period_year=year,
                period_month=month,
                period=period,
                listing_url=listing_url,
                report_page_url="",
                report_title_detected="",
                excel_file_url="",
                source_method="",
                html_fallback_status="",
                html_fallback_path="",
                local_file_path="",
                status="no_report_pages",
                error_message="",
                collected_at=now,
            )
        )
    return rows


def write_report(rows: list[CollectionRow]) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with REPORT_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "period_year",
                "period_month",
                "period",
                "listing_url",
                "report_page_url",
                "report_title_detected",
                "excel_file_url",
                "source_method",
                "html_fallback_status",
                "html_fallback_path",
                "local_file_path",
                "status",
                "error_message",
                "collected_at",
            ]
        )
        for row in rows:
            writer.writerow(
                [
                    row.period_year,
                    row.period_month,
                    row.period,
                    row.listing_url,
                    row.report_page_url,
                    row.report_title_detected,
                    row.excel_file_url,
                    row.source_method,
                    row.html_fallback_status,
                    row.html_fallback_path,
                    row.local_file_path,
                    row.status,
                    row.error_message,
                    row.collected_at,
                ]
            )


def write_summary(rows: list[CollectionRow], started_at: str, finished_at: str) -> None:
    periods = {(r.period_year, r.period_month) for r in rows}
    report_pages_found = {r.report_page_url for r in rows if r.report_page_url}
    report_pages_visited = {r.report_page_url for r in rows if r.report_page_url and r.status != "listing_only"}
    excel_found = sum(1 for r in rows if r.excel_file_url or r.status == "no_excel_found")
    downloaded = sum(1 for r in rows if r.status == "downloaded")
    skipped = sum(1 for r in rows if r.status == "skipped_existing")
    errors = sum(1 for r in rows if r.status == "error")
    html_fallbacks_created = sum(1 for r in rows if r.status == "html_fallback_created")
    html_fallbacks_skipped_existing = sum(1 for r in rows if r.status == "html_fallback_skipped_existing")

    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    with SUMMARY_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "total_months_scanned",
                "total_report_pages_found",
                "total_report_pages_visited",
                "total_excel_files_found",
                "total_excel_files_downloaded",
                "total_excel_files_skipped_existing",
                "total_errors",
                "total_html_fallbacks_created",
                "total_html_fallbacks_skipped_existing",
                "run_started_at",
                "run_finished_at",
            ]
        )
        writer.writerow([
            len(periods),
            len(report_pages_found),
            len(report_pages_visited),
            excel_found,
            downloaded,
            skipped,
            errors,
            html_fallbacks_created,
            html_fallbacks_skipped_existing,
            started_at,
            finished_at,
        ])


def main() -> None:
    args = parse_args()
    if args.year_end < args.year_start:
        raise SystemExit("--year-end must be >= --year-start")

    started_at = datetime.now(timezone.utc).isoformat()
    tasks: list[tuple[int, int]] = [(y, m) for y in range(args.year_start, args.year_end + 1) for m in range(1, 13)]
    all_rows: list[CollectionRow] = []

    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futures = {pool.submit(collect_period, year, month, args.overwrite): (year, month) for year, month in tasks}
        for future in as_completed(futures):
            all_rows.extend(future.result())

    all_rows.sort(key=lambda r: (r.period_year, r.period_month, r.report_page_url, r.excel_file_url))
    write_report(all_rows)
    finished_at = datetime.now(timezone.utc).isoformat()
    write_summary(all_rows, started_at, finished_at)

    print(f"Collection rows: {len(all_rows)} -> {REPORT_PATH}")
    print(f"Summary report -> {SUMMARY_PATH}")


if __name__ == "__main__":
    main()
