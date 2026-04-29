"""Parse CBU 'Profitability indicators of banking sector' Excel files.

Run:
    python normalize_cbu_profitability_indicators.py
    python normalize_cbu_profitability_indicators.py --overwrite
    python normalize_cbu_profitability_indicators.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_profitability_indicators_master.csv")
QA_OUTPUT = Path("data/master/cbu_profitability_indicators_parse_qa.csv")
FILE_NAME_TOKEN = "profitability-indicators-of-banking-sector"
PERIOD_DIR_RE = re.compile(r"^(\d{4})_(\d{2})$")


@dataclass
class MasterRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    indicator: str
    metric_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QaRow:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU profitability indicator reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite outputs if they exist.")
    parser.add_argument(
        "--periods",
        type=str,
        default="",
        help="Comma-separated YYYY_MM folders to parse, e.g. 2026_04,2024_04",
    )
    return parser.parse_args()


def ensure_outputs_writable(overwrite: bool) -> None:
    existing = [path for path in (MASTER_OUTPUT, QA_OUTPUT) if path.exists()]
    if existing and not overwrite:
        names = ", ".join(str(p) for p in existing)
        raise SystemExit(f"Output file(s) already exist: {names}. Use --overwrite to replace them.")


def parse_periods_arg(periods_arg: str) -> Optional[set[str]]:
    if not periods_arg.strip():
        return None
    requested: set[str] = set()
    for token in periods_arg.split(","):
        period = token.strip()
        if not period:
            continue
        match = PERIOD_DIR_RE.match(period)
        if not match:
            raise SystemExit(f"Invalid period '{period}'. Expected YYYY_MM.")
        month = int(match.group(2))
        if not 1 <= month <= 12:
            raise SystemExit(f"Invalid period '{period}'. Month must be 01..12.")
        requested.add(period)
    if not requested:
        raise SystemExit("--periods was provided but no valid values were found.")
    return requested


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    compact = text.replace(" ", "")
    if "," in compact and "." in compact:
        compact = compact.replace(",", "")
    elif "," in compact:
        compact = compact.replace(",", ".")

    try:
        return float(compact)
    except ValueError:
        return None


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def period_label(dt: date) -> str:
    return f"{dt.strftime('%B')} {dt.day}, {dt.year}"


def find_report_period_folder(file_path: Path) -> Optional[str]:
    for parent in file_path.parents:
        if PERIOD_DIR_RE.match(parent.name):
            return parent.name
    return None


def find_input_files(root: Path, requested_periods: Optional[set[str]]) -> list[Path]:
    matches: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            continue
        if FILE_NAME_TOKEN not in path.name.lower():
            continue
        folder = find_report_period_folder(path)
        if requested_periods is not None and folder not in requested_periods:
            continue
        matches.append(path)
    return sorted(matches)


def infer_indicator_metadata(indicator_text: str) -> tuple[str, str, bool]:
    unit = "billion UZS"
    text = indicator_text
    lower = indicator_text.lower()

    if "(in billion uzs)" in lower:
        text = re.sub(r"\(in\s+billion\s+uzs\)", "", text, flags=re.IGNORECASE).strip()
        unit = "billion UZS"
    if "(in %)" in lower:
        text = re.sub(r"\(in\s*%\)", "", text, flags=re.IGNORECASE).strip()
        unit = "percent"

    ratio_tokens = ["ratio", "margin", "roa", "roe", "return on assets", "return on equity", "%"]
    is_ratio = any(token in lower for token in ratio_tokens) or unit == "percent"
    if is_ratio:
        unit = "percent"

    return text.strip(), unit, is_ratio


def parse_file(file_path: Path, loaded_at: str) -> tuple[list[MasterRow], QaRow]:
    report_period_folder = find_report_period_folder(file_path)
    if report_period_folder is None:
        return [], QaRow(str(file_path), "unknown", "", 0, 0, "failed", "Could not infer report period folder.")

    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    ws = wb.active if wb.active is not None else wb.worksheets[0]

    b_period = parse_excel_date(ws["B3"].value)
    c_period = parse_excel_date(ws["C3"].value)
    if b_period is None or c_period is None:
        wb.close()
        return [], QaRow(str(file_path), report_period_folder, ws.title, 0, 0, "failed", "Could not parse B3/C3 period headers.")

    rows_read = 0
    out_rows: list[MasterRow] = []

    max_row = ws.max_row
    for row_idx in range(4, max_row + 1):
        indicator_raw = clean_text(ws.cell(row=row_idx, column=1).value)
        if not indicator_raw:
            continue
        if indicator_raw.startswith("*"):
            continue

        lower_ind = indicator_raw.lower()
        if lower_ind.startswith("note") or lower_ind.startswith("source"):
            continue

        rows_read += 1
        indicator_clean, inferred_unit, is_ratio = infer_indicator_metadata(indicator_raw)

        for col_letter, period_dt in (("B", b_period), ("C", c_period), ("D", c_period)):
            col_idx = ord(col_letter) - ord("A") + 1
            cell = ws.cell(row=row_idx, column=col_idx)
            value = parse_number(cell.value)
            if value is None:
                continue

            if col_letter == "D":
                metric_type = "growth"
                unit = "percent"
            else:
                metric_type = "ratio" if is_ratio else "absolute"
                unit = inferred_unit

            out_rows.append(
                MasterRow(
                    report_period_folder=report_period_folder,
                    period_year=period_dt.year,
                    period_month=period_dt.month,
                    period_date=period_dt.isoformat(),
                    period_label=period_label(period_dt),
                    indicator=indicator_clean,
                    metric_type=metric_type,
                    value=value,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=ws.title,
                    source_cell=f"{get_column_letter(col_idx)}{row_idx}",
                    loaded_at=loaded_at,
                )
            )

    wb.close()

    status = "ok" if out_rows else "warning"
    notes = "Parsed rows from column A and values from B/C/D."
    return out_rows, QaRow(str(file_path), report_period_folder, ws.title, rows_read, len(out_rows), status, notes)


def write_csv(path: Path, rows: list[object], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)


def main() -> None:
    args = parse_args()
    ensure_outputs_writable(overwrite=args.overwrite)
    requested_periods = parse_periods_arg(args.periods)

    input_files = find_input_files(RAW_ROOT, requested_periods)
    if not input_files:
        raise SystemExit("No matching input files were found.")

    loaded_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    master_rows: list[MasterRow] = []
    qa_rows: list[QaRow] = []

    for file_path in input_files:
        file_rows, qa_row = parse_file(file_path, loaded_at)
        master_rows.extend(file_rows)
        qa_rows.append(qa_row)

    write_csv(
        MASTER_OUTPUT,
        master_rows,
        [
            "report_period_folder",
            "period_year",
            "period_month",
            "period_date",
            "period_label",
            "indicator",
            "metric_type",
            "value",
            "unit",
            "source_file",
            "source_sheet",
            "source_cell",
            "loaded_at",
        ],
    )
    write_csv(
        QA_OUTPUT,
        qa_rows,
        [
            "source_file",
            "report_period_folder",
            "source_sheet",
            "rows_read",
            "output_rows_created",
            "status",
            "notes",
        ],
    )

    print(f"Parsed {len(input_files)} files into {MASTER_OUTPUT} and {QA_OUTPUT}.")


if __name__ == "__main__":
    main()
