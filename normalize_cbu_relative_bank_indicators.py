#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_PATH = Path("data/master/cbu_relative_bank_indicators_master.csv")
QA_PATH = Path("data/master/cbu_relative_bank_indicators_parse_qa.csv")
TARGET_TOKEN = "Information-on-relative-indicators-of-banks"

INDICATOR_MAP = {
    "C": "assets",
    "D": "liabilities",
    "E": "assets_to_liabilities_ratio",
    "F": "loans",
    "G": "deposits",
    "H": "loans_to_deposits_ratio",
    "I": "assets",
    "J": "liabilities",
    "K": "assets_to_liabilities_ratio",
    "L": "loans",
    "M": "deposits",
    "N": "loans_to_deposits_ratio",
}

UNIT_MAP = {
    "assets": "billion UZS",
    "liabilities": "billion UZS",
    "loans": "billion UZS",
    "deposits": "billion UZS",
    "assets_to_liabilities_ratio": "ratio",
    "loans_to_deposits_ratio": "ratio",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU relative bank indicators Excel reports.")
    parser.add_argument("--periods", help="Comma-separated list of YYYY_MM folders to parse.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing outputs.")
    return parser.parse_args()


def parse_periods_arg(raw: str | None) -> set[str] | None:
    if not raw:
        return None
    periods = {p.strip() for p in raw.split(",") if p.strip()}
    invalid = [p for p in periods if not re.fullmatch(r"\d{4}_\d{2}", p)]
    if invalid:
        raise ValueError(f"Invalid --periods value(s): {', '.join(sorted(invalid))}. Expected YYYY_MM.")
    return periods


def iter_target_files(selected_periods: set[str] | None) -> Iterable[Path]:
    if not RAW_ROOT.exists():
        return []

    matches: list[Path] = []
    for period_dir in sorted(p for p in RAW_ROOT.iterdir() if p.is_dir()):
        if selected_periods is not None and period_dir.name not in selected_periods:
            continue
        for f in sorted(period_dir.glob("*.xlsx")):
            if TARGET_TOKEN in f.name:
                matches.append(f)
    return matches


def excel_date_to_label(value) -> tuple[datetime | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        dt = None
        for fmt in ("%B %d, %Y", "%b %d, %Y", "%d.%m.%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            return None, text
    label = f"{dt.strftime('%B')} {dt.day}, {dt.year}"
    return dt, label


def header_period_labels(ws) -> dict[str, tuple[datetime | None, str | None]]:
    labels: dict[str, tuple[datetime | None, str | None]] = {}
    for col_idx in range(3, 15):
        col = get_column_letter(col_idx)
        labels[col] = excel_date_to_label(ws[f"{col}3"].value)
    return labels


def is_numbered_row(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    text = str(value).strip()
    return bool(re.fullmatch(r"\d+", text))


def to_number_or_none(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_workbook(path: Path) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(path, data_only=True)
    master_rows: list[dict] = []
    qa_rows: list[dict] = []
    report_period_folder = path.parent.name
    now_iso = datetime.utcnow().isoformat(timespec="seconds")

    for ws in wb.worksheets:
        period_map = header_period_labels(ws)
        rows_read = 0
        out_count_start = len(master_rows)
        current_group = "Unknown"

        max_row = ws.max_row
        for r in range(4, max_row + 1):
            col_a = ws[f"A{r}"].value
            col_b = ws[f"B{r}"].value

            if col_a is None and col_b is None:
                continue

            rows_read += 1
            a_text = str(col_a).strip() if col_a is not None else ""
            b_text = str(col_b).strip() if col_b is not None else ""

            bank_group = None
            bank_name = None

            if "Total" in a_text:
                bank_group = "Total"
                bank_name = "Total"
                current_group = "Total"
            elif "Banks with State ownership" in a_text:
                bank_group = "Banks with State ownership"
                bank_name = "Banks with State ownership"
                current_group = "Banks with State ownership"
            elif "Other banks" in a_text:
                bank_group = "Other banks"
                bank_name = "Other banks"
                current_group = "Other banks"
            elif b_text and is_numbered_row(col_a):
                bank_group = current_group
                bank_name = b_text
            else:
                continue

            for col, indicator in INDICATOR_MAP.items():
                val = to_number_or_none(ws[f"{col}{r}"].value)
                if val is None:
                    continue

                dt, label = period_map.get(col, (None, None))
                master_rows.append(
                    {
                        "report_period_folder": report_period_folder,
                        "period_year": dt.year if dt else None,
                        "period_month": dt.month if dt else None,
                        "period_date": dt.date().isoformat() if dt else None,
                        "period_label": label,
                        "bank_group": bank_group,
                        "bank_name": bank_name,
                        "indicator": indicator,
                        "value": val,
                        "unit": UNIT_MAP[indicator],
                        "source_file": str(path),
                        "source_sheet": ws.title,
                        "source_cell": f"{col}{r}",
                        "loaded_at": now_iso,
                    }
                )

        qa_rows.append(
            {
                "source_file": str(path),
                "report_period_folder": report_period_folder,
                "source_sheet": ws.title,
                "rows_read": rows_read,
                "output_rows_created": len(master_rows) - out_count_start,
                "status": "ok",
                "notes": "",
            }
        )

    return master_rows, qa_rows


def main() -> int:
    args = parse_args()
    try:
        selected_periods = parse_periods_arg(args.periods)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 2

    if (MASTER_PATH.exists() or QA_PATH.exists()) and not args.overwrite:
        print(
            f"Output already exists. Use --overwrite to replace:\n- {MASTER_PATH}\n- {QA_PATH}",
            file=sys.stderr,
        )
        return 1

    files = list(iter_target_files(selected_periods))
    all_master: list[dict] = []
    all_qa: list[dict] = []

    if not files:
        print("No matching files found to parse.")

    for file_path in files:
        master_rows, qa_rows = parse_workbook(file_path)
        all_master.extend(master_rows)
        all_qa.extend(qa_rows)

    MASTER_PATH.parent.mkdir(parents=True, exist_ok=True)

    pd.DataFrame(
        all_master,
        columns=[
            "report_period_folder",
            "period_year",
            "period_month",
            "period_date",
            "period_label",
            "bank_group",
            "bank_name",
            "indicator",
            "value",
            "unit",
            "source_file",
            "source_sheet",
            "source_cell",
            "loaded_at",
        ],
    ).to_csv(MASTER_PATH, index=False)

    pd.DataFrame(
        all_qa,
        columns=[
            "source_file",
            "report_period_folder",
            "source_sheet",
            "rows_read",
            "output_rows_created",
            "status",
            "notes",
        ],
    ).to_csv(QA_PATH, index=False)

    print(f"Wrote master: {MASTER_PATH}")
    print(f"Wrote QA: {QA_PATH}")
    print(f"Files parsed: {len(files)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
