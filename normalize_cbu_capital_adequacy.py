"""Parse CBU "Capital adequacy of the banking sector" Excel files into master + QA CSV outputs.

Run:
    python normalize_cbu_capital_adequacy.py
    python normalize_cbu_capital_adequacy.py --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_capital_adequacy_master.csv")
QA_OUTPUT = Path("data/master/cbu_capital_adequacy_parse_qa.csv")
FILE_NAME_CORE_PHRASE = "capital-adequacy-of-the-banking-sector"
DATE_HEADER_ROW = 3
UNIT_ROW = 4
DATA_START_ROW = 5


@dataclass
class ParsedRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    indicator: str
    metric_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QAEntry:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU capital adequacy Excel reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files if they already exist.")
    parser.add_argument(
        "--periods",
        type=str,
        help="Optional comma-separated YYYY_MM folders to parse (for example: 2026_04,2024_04).",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return " ".join(text.split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if not text:
        return None

    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def parse_periods_arg(periods_raw: Optional[str]) -> list[str]:
    if periods_raw is None:
        return []

    periods = [part.strip() for part in periods_raw.split(",") if part.strip()]
    period_pattern = re.compile(r"^\d{4}_\d{2}$")
    invalid = [period for period in periods if not period_pattern.match(period)]
    if invalid:
        raise SystemExit(
            "Invalid --periods value(s): "
            f"{', '.join(invalid)}. Expected comma-separated YYYY_MM values."
        )
    return periods


def find_input_files(root_dir: Path, allowed_periods: Optional[set[str]] = None) -> list[Path]:
    files: list[Path] = []
    for path in root_dir.rglob("*.xlsx"):
        if FILE_NAME_CORE_PHRASE in path.name.lower():
            if allowed_periods is not None:
                _, _, period_folder = infer_period_folder(path)
                if period_folder not in allowed_periods:
                    continue
            files.append(path)
    return sorted(files)


def infer_period_folder(file_path: Path) -> tuple[Optional[int], Optional[int], str]:
    pattern = re.compile(r"^(\d{4})_(\d{2})$")
    for parent in file_path.parents:
        m = pattern.match(parent.name)
        if m:
            y = int(m.group(1))
            mo = int(m.group(2))
            if 1 <= mo <= 12:
                return y, mo, parent.name
    return None, None, "unknown"


def build_merged_lookup(sheet: Worksheet) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for merged in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = sheet.cell(row=min_row, column=min_col).coordinate
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                coord = sheet.cell(row=r, column=c).coordinate
                lookup[coord] = top_left
    return lookup


def get_effective_value(sheet: Worksheet, cell_ref: str, merged_lookup: dict[str, str]) -> tuple[object, str]:
    source_cell = merged_lookup.get(cell_ref, cell_ref)
    return sheet[source_cell].value, source_cell


def parse_date_header(raw: object) -> tuple[str, str]:
    if raw is None:
        return "", ""
    if isinstance(raw, (datetime, date)):
        dt = raw.date() if isinstance(raw, datetime) else raw
        return dt.isoformat(), dt.strftime("%B %-d, %Y")
    text = clean_text(raw)
    if not text:
        return "", ""

    for fmt in ("%B %d, %Y", "%B %d %Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(text, fmt).date()
            return dt.isoformat(), dt.strftime("%B %-d, %Y")
        except ValueError:
            continue
    return "", text


def detect_metric_type_and_unit(unit_text: str, indicator_text: str) -> tuple[str, str]:
    unit_norm = unit_text.lower()
    ind_norm = indicator_text.lower()
    ratio_cues = ["%", "percent", "ratio", "coefficient", "times", "x"]

    if any(cue in unit_norm for cue in ratio_cues) or any(cue in ind_norm for cue in ["ratio", "coefficient", "%"]):
        return "ratio", "ratio"
    if unit_text:
        return "absolute", unit_text
    return "unknown", "unknown"


def parse_workbook(file_path: Path, loaded_at: str) -> tuple[list[ParsedRow], QAEntry]:
    year, month, period_folder = infer_period_folder(file_path)
    if year is None or month is None:
        return [], QAEntry(str(file_path), period_folder, "", 0, 0, "failed", "Could not infer YYYY_MM folder from path.")

    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active if wb.active else wb.worksheets[0]
    merged_lookup = build_merged_lookup(sheet)

    output: list[ParsedRow] = []
    notes: list[str] = []
    rows_read = 0

    max_col = sheet.max_column
    max_row = sheet.max_row

    for row_idx in range(DATA_START_ROW, max_row + 1):
        indicator = clean_text(sheet[f"A{row_idx}"].value)
        if not indicator:
            continue

        row_created = 0
        for col_idx in range(2, max_col + 1):
            col_letter = sheet.cell(row=1, column=col_idx).column_letter
            value_ref = f"{col_letter}{row_idx}"
            raw_value, source_cell = get_effective_value(sheet, value_ref, merged_lookup)
            num = parse_number(raw_value)
            if num is None:
                continue

            date_ref = f"{col_letter}{DATE_HEADER_ROW}"
            date_raw, _ = get_effective_value(sheet, date_ref, merged_lookup)
            period_date, period_label = parse_date_header(date_raw)
            if not period_label:
                period_label = f"column_{col_letter}"
                notes.append(f"Missing date header at {date_ref}; used fallback label.")
            if not period_date:
                period_date = f"{year:04d}-{month:02d}-01"

            unit_ref = f"{col_letter}{UNIT_ROW}"
            unit_raw, _ = get_effective_value(sheet, unit_ref, merged_lookup)
            unit_text = clean_text(unit_raw)

            metric_type, unit = detect_metric_type_and_unit(unit_text, indicator)

            output.append(
                ParsedRow(
                    report_period_folder=period_folder,
                    period_year=year,
                    period_month=month,
                    period_date=period_date,
                    period_label=period_label,
                    indicator=indicator,
                    metric_type=metric_type,
                    value=num,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=source_cell,
                    loaded_at=loaded_at,
                )
            )
            row_created += 1

        if row_created == 0:
            notes.append(f"Indicator row A{row_idx} had no numeric cells in B:{sheet.cell(row=1, column=max_col).column_letter}.")
        else:
            rows_read += 1

    status = "ok" if output else "warning"
    note_text = "Parsed successfully." if not notes else " | ".join(dict.fromkeys(notes))
    qa = QAEntry(str(file_path), period_folder, sheet.title, rows_read, len(output), status, note_text)
    return output, qa


def write_master(rows: list[ParsedRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "report_period_folder", "period_year", "period_month", "period_date", "period_label", "indicator",
            "metric_type", "value", "unit", "source_file", "source_sheet", "source_cell", "loaded_at"
        ])
        for r in rows:
            w.writerow([
                r.report_period_folder, r.period_year, r.period_month, r.period_date, r.period_label, r.indicator,
                r.metric_type, r.value, r.unit, r.source_file, r.source_sheet, r.source_cell, r.loaded_at
            ])


def write_qa(rows: list[QAEntry], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["source_file", "report_period_folder", "source_sheet", "rows_read", "output_rows_created", "status", "notes"])
        for r in rows:
            w.writerow([r.source_file, r.report_period_folder, r.source_sheet, r.rows_read, r.output_rows_created, r.status, r.notes])


def main() -> None:
    args = parse_args()
    if not args.overwrite and (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()):
        raise SystemExit(
            "Output file already exists. Re-run with --overwrite to replace: "
            f"{MASTER_OUTPUT} / {QA_OUTPUT}"
        )

    periods = parse_periods_arg(args.periods)
    period_filter = set(periods) if periods else None
    files = find_input_files(RAW_ROOT, period_filter)
    if not files:
        if periods:
            raise SystemExit(
                "No input files found for requested --periods under "
                f"{RAW_ROOT} matching '*{FILE_NAME_CORE_PHRASE}*.xlsx'."
            )
        raise SystemExit(f"No input files found under {RAW_ROOT} matching '*{FILE_NAME_CORE_PHRASE}*.xlsx'.")

    loaded_at = datetime.now(timezone.utc).isoformat()
    all_rows: list[ParsedRow] = []
    qa_rows: list[QAEntry] = []

    if periods:
        for period in periods:
            period_path = RAW_ROOT / period
            if not period_path.exists():
                warning_note = f"Period folder does not exist: {period_path}"
                print(f"WARNING: {warning_note}")
                qa_rows.append(
                    QAEntry(
                        source_file="",
                        report_period_folder=period,
                        source_sheet="",
                        rows_read=0,
                        output_rows_created=0,
                        status="warning",
                        notes=warning_note,
                    )
                )

    for file_path in files:
        parsed, qa = parse_workbook(file_path, loaded_at)
        all_rows.extend(parsed)
        qa_rows.append(qa)

    write_master(all_rows, MASTER_OUTPUT)
    write_qa(qa_rows, QA_OUTPUT)
    print(f"Parsed {len(files)} files -> {len(all_rows)} rows")
    print(f"Master: {MASTER_OUTPUT}")
    print(f"QA: {QA_OUTPUT}")


if __name__ == "__main__":
    main()
