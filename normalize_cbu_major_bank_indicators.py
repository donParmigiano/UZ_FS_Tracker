"""Normalize recurring CBU "Information on major indicators of commercial banks" Excel reports.

This script scans downloaded raw CBU files under data/raw/cbu_bankstats/YYYY_MM/
and builds one analytical master table from files whose names contain
"Information-on-major-indicators-of-commercial-banks".
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_major_bank_indicators_master.csv")
QA_OUTPUT = Path("data/master/cbu_major_bank_indicators_parse_qa.csv")
TARGET_NAME_TOKEN = "information-on-major-indicators-of-commercial-banks"

MASTER_COLUMNS = [
    "period_year",
    "period_month",
    "period",
    "bank_group",
    "bank_name",
    "indicator",
    "metric_type",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "source_cell",
    "loaded_at",
]

QA_COLUMNS = [
    "source_file",
    "period",
    "rows_read",
    "output_rows_created",
    "status",
    "notes",
]

GROUP_KEYWORDS = [
    "state",
    "other",
    "group",
    "total",
    "jami",
    "davlat",
    "boshqa",
    "акционер",
    "группа",
    "итого",
]

INDICATOR_KEYWORDS = {
    "assets": ["asset", "assets", "актив", "активы"],
    "loans": ["loan", "loans", "credit", "кредит", "ссуда"],
    "capital": ["capital", "капитал"],
    "deposits": ["deposit", "deposits", "депозит", "вклад"],
}


@dataclass
class MetricColumn:
    """Metadata for one indicator+metric column in the report."""

    col_idx: int
    indicator: str
    metric_type: str
    unit: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize recurring CBU major bank indicators report")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing master/QA outputs")
    return parser.parse_args()


def discover_target_files(root: Path) -> list[Path]:
    """Find raw files that match the recurring report naming pattern."""
    candidates: list[Path] = []
    for path in root.glob("*/*"):
        if not path.is_file():
            continue
        lowered = path.name.lower()
        if TARGET_NAME_TOKEN in lowered and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            candidates.append(path)
    return sorted(candidates)


def period_from_path(path: Path) -> tuple[str, str, str]:
    """Read YYYY_MM period from the parent folder name."""
    match = re.match(r"(\d{4})_(\d{2})", path.parent.name)
    if not match:
        return "", "", ""
    year, month = match.group(1), match.group(2)
    return year, month, f"{year}-{month}"


def as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def is_numeric(value: object) -> bool:
    return isinstance(value, (int, float))


def looks_like_percent(text: str) -> bool:
    t = text.lower()
    return "%" in t or "share" in t or "ulush" in t or "доля" in t


def detect_indicator(header_text: str) -> str:
    text = header_text.lower()
    for indicator, keywords in INDICATOR_KEYWORDS.items():
        if any(keyword in text for keyword in keywords):
            return indicator
    return "needs_review"


def detect_metric_columns(sheet) -> list[MetricColumn]:
    """Infer indicator/metric columns from merged and multi-row headers.

    We read header text from first rows and classify each data column as:
    - absolute value (unit defaults to billion_uzs)
    - share/percentage (unit percent)
    """
    metric_columns: list[MetricColumn] = []

    # Combine text from first 8 rows for each column to capture merged headers.
    for col_idx in range(2, sheet.max_column + 1):
        header_parts: list[str] = []
        for row_idx in range(1, min(8, sheet.max_row) + 1):
            header_parts.append(as_text(sheet.cell(row=row_idx, column=col_idx).value))
        header_text = normalize_space(" ".join(part for part in header_parts if part))
        if not header_text:
            continue

        indicator = detect_indicator(header_text)
        if indicator == "needs_review":
            continue

        metric_type = "share" if looks_like_percent(header_text) else "absolute"
        unit = "percent" if metric_type == "share" else "billion_uzs"
        metric_columns.append(MetricColumn(col_idx=col_idx, indicator=indicator, metric_type=metric_type, unit=unit))

    # De-duplicate columns while preserving order.
    seen: set[int] = set()
    deduped: list[MetricColumn] = []
    for col in metric_columns:
        if col.col_idx not in seen:
            seen.add(col.col_idx)
            deduped.append(col)
    return deduped


def first_text_cell(row_values: Iterable[object]) -> str:
    for value in row_values:
        text = normalize_space(as_text(value))
        if text:
            return text
    return ""


def classify_row(label: str, current_group: str) -> tuple[str, str, bool]:
    """Return (bank_group, bank_name, is_group_header)."""
    clean_label = normalize_space(label)
    if not clean_label:
        return current_group or "needs_review", "needs_review", False

    lowered = clean_label.lower()
    if any(keyword in lowered for keyword in GROUP_KEYWORDS):
        # Group rows can be totals or category headers.
        return clean_label, clean_label, True

    if current_group:
        return current_group, clean_label, False

    return "needs_review", clean_label, False


def parse_workbook(path: Path, loaded_at: str) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Parse one workbook and emit normalized rows + one QA row."""
    year, month, period = period_from_path(path)
    output_rows: list[dict[str, object]] = []
    rows_read = 0
    notes: list[str] = []

    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        qa = {
            "source_file": str(path),
            "period": period,
            "rows_read": 0,
            "output_rows_created": 0,
            "status": "failed",
            "notes": f"workbook_open_error: {exc}",
        }
        return output_rows, qa

    for sheet in workbook.worksheets:
        metric_cols = detect_metric_columns(sheet)
        if not metric_cols:
            notes.append(f"{sheet.title}: no metric columns detected")
            continue

        current_group = ""
        # Data rows usually begin after multi-row headers.
        for row_idx in range(4, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
            label = first_text_cell(row_values[:2])

            # Skip empty lines with no numeric content.
            has_numeric = any(is_numeric(v) for v in row_values)
            if not label and not has_numeric:
                continue

            rows_read += 1
            bank_group, bank_name, is_group_header = classify_row(label, current_group)
            if is_group_header:
                current_group = bank_group

            for metric in metric_cols:
                cell = sheet.cell(row=row_idx, column=metric.col_idx)
                if not is_numeric(cell.value):
                    continue

                output_rows.append(
                    {
                        "period_year": year,
                        "period_month": month,
                        "period": period,
                        "bank_group": bank_group,
                        "bank_name": bank_name,
                        "indicator": metric.indicator,
                        "metric_type": metric.metric_type,
                        "value": float(cell.value),
                        "unit": metric.unit,
                        "source_file": str(path),
                        "source_sheet": sheet.title,
                        "source_cell": cell.coordinate,
                        "loaded_at": loaded_at,
                    }
                )

    status = "ok" if output_rows else "warning"
    qa = {
        "source_file": str(path),
        "period": period,
        "rows_read": rows_read,
        "output_rows_created": len(output_rows),
        "status": status,
        "notes": "; ".join(notes) if notes else "",
    }
    return output_rows, qa


def write_csv(path: Path, columns: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def run(overwrite: bool) -> None:
    if not overwrite and MASTER_OUTPUT.exists():
        raise SystemExit(f"{MASTER_OUTPUT} exists. Re-run with --overwrite to replace it.")
    if not overwrite and QA_OUTPUT.exists():
        raise SystemExit(f"{QA_OUTPUT} exists. Re-run with --overwrite to replace it.")

    target_files = discover_target_files(RAW_ROOT)
    loaded_at = datetime.now(timezone.utc).isoformat()

    all_rows: list[dict[str, object]] = []
    qa_rows: list[dict[str, object]] = []

    if not target_files:
        qa_rows.append(
            {
                "source_file": "",
                "period": "",
                "rows_read": 0,
                "output_rows_created": 0,
                "status": "warning",
                "notes": "No target files found under data/raw/cbu_bankstats",
            }
        )
    else:
        for path in target_files:
            rows, qa = parse_workbook(path, loaded_at)
            all_rows.extend(rows)
            qa_rows.append(qa)

    write_csv(MASTER_OUTPUT, MASTER_COLUMNS, all_rows)
    write_csv(QA_OUTPUT, QA_COLUMNS, qa_rows)

    print(f"Done. files={len(target_files)} output_rows={len(all_rows)} qa_rows={len(qa_rows)}")


if __name__ == "__main__":
    args = parse_args()
    run(overwrite=args.overwrite)
