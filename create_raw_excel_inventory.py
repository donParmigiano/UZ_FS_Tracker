from pathlib import Path
from openpyxl import load_workbook
import csv

ROOT = Path(r"C:\Users\skadyrbekov\OneDrive - KPMG\Desktop\UZ_FS_Tracker_git\data\raw\cbu_bankstats")
OUT = Path(r"C:\Users\skadyrbekov\OneDrive - KPMG\Desktop\UZ_FS_Tracker_git\data\processed\raw_excel_inventory.csv")

rows = []

for file_path in sorted(ROOT.rglob("*")):
    if file_path.suffix.lower() not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        continue
    if file_path.name.startswith("~$"):
        continue

    period_folder = file_path.parent.name
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows.append({
                "period_folder": period_folder,
                "file_name": file_path.name,
                "file_path": str(file_path),
                "file_size_bytes": file_path.stat().st_size,
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "status": "ok",
                "error": "",
            })
        wb.close()
    except Exception as e:
        rows.append({
            "period_folder": period_folder,
            "file_name": file_path.name,
            "file_path": str(file_path),
            "file_size_bytes": file_path.stat().st_size,
            "sheet_name": "",
            "max_row": "",
            "max_column": "",
            "status": "error",
            "error": str(e),
        })

OUT.parent.mkdir(parents=True, exist_ok=True)

with OUT.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "period_folder",
        "file_name",
        "file_path",
        "file_size_bytes",
        "sheet_name",
        "max_row",
        "max_column",
        "status",
        "error",
    ])
    writer.writeheader()
    writer.writerows(rows)

print(f"Inventory created: {OUT}")
print(f"Rows: {len(rows)}")
