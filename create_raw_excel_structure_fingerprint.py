from pathlib import Path
from openpyxl import load_workbook
import csv

ROOT = Path(r"C:\Users\skadyrbekov\OneDrive - KPMG\Desktop\UZ_FS_Tracker_git\data\raw\cbu_bankstats")
OUT = Path(r"C:\Users\skadyrbekov\OneDrive - KPMG\Desktop\UZ_FS_Tracker_git\data\processed\raw_excel_structure_fingerprint.csv")

MAX_ROWS = 15
MAX_COLS = 15

rows = []

for file_path in sorted(ROOT.rglob("*")):
    if file_path.suffix.lower() not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        continue
    if file_path.name.startswith("~$"):
        continue

    period_folder = file_path.parent.name

    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        for ws in wb.worksheets:
            for r in range(1, min(ws.max_row, MAX_ROWS) + 1):
                for c in range(1, min(ws.max_column, MAX_COLS) + 1):
                    value = ws.cell(row=r, column=c).value
                    if value is not None and str(value).strip() != "":
                        rows.append({
                            "period_folder": period_folder,
                            "file_name": file_path.name,
                            "sheet_name": ws.title,
                            "row": r,
                            "column": c,
                            "cell_value": str(value).strip(),
                        })
        wb.close()
    except Exception as e:
        rows.append({
            "period_folder": period_folder,
            "file_name": file_path.name,
            "sheet_name": "",
            "row": "",
            "column": "",
            "cell_value": f"ERROR: {e}",
        })

OUT.parent.mkdir(parents=True, exist_ok=True)

with OUT.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "period_folder",
        "file_name",
        "sheet_name",
        "row",
        "column",
        "cell_value",
    ])
    writer.writeheader()
    writer.writerows(rows)

print(f"Fingerprint created: {OUT}")
print(f"Rows: {len(rows)}")
