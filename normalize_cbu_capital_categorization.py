"""Parse CBU "Categorization of commercial banks in terms of total and authorized capital levels" reports.

Run:
    python normalize_cbu_capital_categorization.py
    python normalize_cbu_capital_categorization.py --overwrite
    python normalize_cbu_capital_categorization.py --periods 2026_04 --overwrite
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUTPUT = Path("data/master/cbu_capital_categorization_master.csv")
QA_OUTPUT = Path("data/master/cbu_capital_categorization_parse_qa.csv")
FILE_NAME_CORE_PHRASE = "categorization-of-commercial-banks-in-terms-of-total-and-authorized-capital-levels"

COLUMN_MAP = {
    2: ("total", "number_of_banks", "banks"),
    3: ("total", "absolute_value", "billion UZS"),
    4: ("up to 500 billion UZS", "number_of_banks", "banks"),
    5: ("up to 500 billion UZS", "share", "percent"),
    6: ("from 500 billion to 1 trillion UZS", "number_of_banks", "banks"),
    7: ("from 500 billion to 1 trillion UZS", "share", "percent"),
    8: ("from 1 to 2 trillion UZS", "number_of_banks", "banks"),
    9: ("from 1 to 2 trillion UZS", "share", "percent"),
    10: ("from 2 to 5 trillion UZS", "number_of_banks", "banks"),
    11: ("from 2 to 5 trillion UZS", "share", "percent"),
    12: ("from 5 to 10 trillion UZS", "number_of_banks", "banks"),
    13: ("from 5 to 10 trillion UZS", "share", "percent"),
    14: ("Above 10 trillion UZS", "number_of_banks", "banks"),
    15: ("Above 10 trillion UZS", "share", "percent"),
}


@dataclass
class ParsedRow:
    report_period_folder: str
    period_year: int
    period_month: int
    period_date: str
    period_label: str
    indicator: str
    capital_category: str
    metric_type: str
    value: float
    unit: str
    source_file: str
    source_sheet: str
    source_cell: str
    loaded_at: str


@dataclass
class QAEntry:
    source_file: str
    report_period_folder: str
    source_sheet: str
    rows_read: int
    output_rows_created: int
    status: str
    notes: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU capital categorization Excel reports.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite output CSVs if they already exist.")
    parser.add_argument(
        "--periods",
        type=str,
        help="Comma-separated YYYY_MM folders to parse under data/raw/cbu_bankstats (e.g. 2026_04,2024_04).",
    )
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return " ".join(text.split())


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    compact = text.replace(" ", "")
    if "," in compact and "." in compact:
        compact = compact.replace(",", "")
    elif "," in compact:
        compact = compact.replace(",", ".")
    try:
        return float(compact)
    except ValueError:
        return None


def parse_periods_arg(periods_arg: Optional[str]) -> list[str]:
    if not periods_arg:
        return []
    periods = [part.strip() for part in periods_arg.split(",") if part.strip()]
    pattern = re.compile(r"^\d{4}_(0[1-9]|1[0-2])$")
    invalid = [p for p in periods if not pattern.match(p)]
    if invalid:
        raise SystemExit(
            "Invalid --periods value(s): "
            + ", ".join(invalid)
            + ". Expected comma-separated YYYY_MM values (e.g. 2026_04,2024_04)."
        )
    return periods


def find_input_files(root_dir: Path, periods_filter: list[str]) -> list[Path]:
    period_set = set(periods_filter)
    files: list[Path] = []
    for path in root_dir.rglob("*.xlsx"):
        if FILE_NAME_CORE_PHRASE not in path.name.lower():
            continue
        if period_set:
            folder = infer_period_folder(path)[2]
            if folder not in period_set:
                continue
        files.append(path)
    return sorted(files)


def infer_period_folder(file_path: Path) -> tuple[Optional[int], Optional[int], str]:
    pattern = re.compile(r"^(\d{4})_(\d{2})$")
    for parent in file_path.parents:
        match = pattern.match(parent.name)
        if not match:
            continue
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return year, month, parent.name
    return None, None, "unknown"


def build_merged_lookup(sheet: Worksheet) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for merged in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = sheet.cell(row=min_row, column=min_col).coordinate
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[sheet.cell(row=row, column=col).coordinate] = top_left
    return lookup


def get_effective_cell(sheet: Worksheet, row: int, col: int, merged_lookup: dict[str, str]) -> tuple[object, str]:
    ref = sheet.cell(row=row, column=col).coordinate
    source_ref = merged_lookup.get(ref, ref)
    return sheet[source_ref].value, source_ref


def parse_period_label_and_date(value: object) -> tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    norm = text.lower()
    if norm.startswith("as of"):
        norm = norm[5:].strip()
    norm = norm.replace("sept", "sep")
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(norm, fmt).date()
            return text, parsed.isoformat()
        except ValueError:
            continue
    return text, ""


def parse_workbook(file_path: Path, loaded_at: str) -> tuple[list[ParsedRow], QAEntry]:
    year, month, period_folder = infer_period_folder(file_path)
    if year is None or month is None:
        return [], QAEntry(str(file_path), period_folder, "", 0, 0, "failed", "Could not infer YYYY_MM folder from path.")

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active if workbook.active else workbook.worksheets[0]
    merged_lookup = build_merged_lookup(sheet)

    output: list[ParsedRow] = []
    notes: list[str] = []
    rows_read = 0
    current_label = ""
    current_date = ""

    for row in range(1, sheet.max_row + 1):
        col_a_value, col_a_source = get_effective_cell(sheet, row, 1, merged_lookup)
        col_a_text = clean_text(col_a_value)
        if not col_a_text:
            continue

        merged_across_a_to_o = False
        if col_a_source in sheet.merged_cells:
            merged_range = next((r for r in sheet.merged_cells.ranges if col_a_source in r), None)
            if merged_range and merged_range.min_col == 1 and merged_range.max_col >= 15:
                merged_across_a_to_o = True

        period_label, period_date = parse_period_label_and_date(col_a_text)
        if merged_across_a_to_o and period_label:
            current_label = period_label
            current_date = period_date or f"{year:04d}-{month:02d}-01"
            rows_read += 1
            continue

        if not current_label:
            notes.append(f"Skipped row {row}: indicator before first period header.")
            rows_read += 1
            continue

        indicator = col_a_text
        rows_read += 1
        for col, (category, metric_type, unit) in COLUMN_MAP.items():
            raw_value, source_cell = get_effective_cell(sheet, row, col, merged_lookup)
            num = parse_number(raw_value)
            if num is None:
                continue
            output.append(
                ParsedRow(
                    report_period_folder=period_folder,
                    period_year=year,
                    period_month=month,
                    period_date=current_date,
                    period_label=current_label,
                    indicator=indicator,
                    capital_category=category,
                    metric_type=metric_type,
                    value=num,
                    unit=unit,
                    source_file=str(file_path),
                    source_sheet=sheet.title,
                    source_cell=source_cell,
                    loaded_at=loaded_at,
                )
            )

    status = "ok" if output else "warning"
    notes_text = " | ".join(dict.fromkeys(notes)) if notes else ""
    return output, QAEntry(str(file_path), period_folder, sheet.title, rows_read, len(output), status, notes_text)


def write_master(rows: list[ParsedRow], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "report_period_folder",
                "period_year",
                "period_month",
                "period_date",
                "period_label",
                "indicator",
                "capital_category",
                "metric_type",
                "value",
                "unit",
                "source_file",
                "source_sheet",
                "source_cell",
                "loaded_at",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)


def write_qa(entries: list[QAEntry], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "source_file",
                "report_period_folder",
                "source_sheet",
                "rows_read",
                "output_rows_created",
                "status",
                "notes",
            ],
        )
        writer.writeheader()
        for entry in entries:
            writer.writerow(entry.__dict__)


def main() -> None:
    args = parse_args()
    if (MASTER_OUTPUT.exists() or QA_OUTPUT.exists()) and not args.overwrite:
        raise SystemExit(
            f"Output file exists. Use --overwrite to replace existing outputs: {MASTER_OUTPUT} and/or {QA_OUTPUT}."
        )

    periods = parse_periods_arg(args.periods)
    files = find_input_files(RAW_ROOT, periods)
    if not files:
        period_note = f" for periods {','.join(periods)}" if periods else ""
        raise SystemExit(f"No input files found matching phrase '{FILE_NAME_CORE_PHRASE}'{period_note}.")

    loaded_at = datetime.now(timezone.utc).isoformat()
    master_rows: list[ParsedRow] = []
    qa_rows: list[QAEntry] = []

    for file_path in files:
        rows, qa = parse_workbook(file_path, loaded_at)
        master_rows.extend(rows)
        qa_rows.append(qa)

    write_master(master_rows, MASTER_OUTPUT)
    write_qa(qa_rows, QA_OUTPUT)

    print(f"Parsed {len(files)} file(s).")
    print(f"Master rows: {len(master_rows)} -> {MASTER_OUTPUT}")
    print(f"QA rows: {len(qa_rows)} -> {QA_OUTPUT}")


if __name__ == "__main__":
    main()
