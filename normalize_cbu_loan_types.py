#!/usr/bin/env python3
"""Normalize CBU loan types report files into master and QA CSV outputs."""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

RAW_ROOT = Path("data/raw/cbu_bankstats")
MASTER_OUT = Path("data/master/cbu_loan_types_master.csv")
QA_OUT = Path("data/master/cbu_loan_types_parse_qa.csv")
TARGET_TOKEN = "Information-on-the-loan-portfolio-in-terms-of-loan-types"

MASTER_COLUMNS = [
    "report_period_folder",
    "period_year",
    "period_month",
    "period_date",
    "period_label",
    "loan_type",
    "borrower_type",
    "metric_type",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "source_cell",
    "loaded_at",
]

QA_COLUMNS = [
    "source_file",
    "report_period_folder",
    "source_sheet",
    "rows_read",
    "output_rows_created",
    "status",
    "notes",
]


@dataclass
class PeriodInfo:
    year: int
    month: int
    dt: date


PERIOD_RE = re.compile(r"^\d{4}_\d{2}$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse CBU loan types report files.")
    parser.add_argument(
        "--periods",
        help="Comma-separated YYYY_MM folders under data/raw/cbu_bankstats",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite existing output files.",
    )
    return parser.parse_args()


def ensure_outputs_writable(overwrite: bool) -> None:
    if overwrite:
        return
    existing = [str(p) for p in (MASTER_OUT, QA_OUT) if p.exists()]
    if existing:
        raise SystemExit(
            "Output file(s) already exist. Use --overwrite to replace: " + ", ".join(existing)
        )


def parse_folder_period(folder: str) -> Optional[PeriodInfo]:
    if not PERIOD_RE.match(folder):
        return None
    year, month = folder.split("_")
    try:
        dt = date(int(year), int(month), 1)
    except ValueError:
        return None
    return PeriodInfo(year=dt.year, month=dt.month, dt=dt)


def list_target_files(periods_arg: Optional[str]) -> List[Tuple[Path, str]]:
    folders: List[Path] = []
    if periods_arg:
        requested = [p.strip() for p in periods_arg.split(",") if p.strip()]
        bad = [p for p in requested if not PERIOD_RE.match(p)]
        if bad:
            raise SystemExit(f"Invalid --periods values (expected YYYY_MM): {', '.join(bad)}")
        folders = [RAW_ROOT / p for p in requested]
    else:
        if not RAW_ROOT.exists():
            return []
        folders = [p for p in RAW_ROOT.iterdir() if p.is_dir()]

    targets: List[Tuple[Path, str]] = []
    for folder in sorted(folders):
        if not folder.exists():
            continue
        folder_name = folder.name
        for file_path in sorted(folder.iterdir()):
            if not file_path.is_file():
                continue
            if TARGET_TOKEN not in file_path.name:
                continue
            if file_path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                continue
            targets.append((file_path, folder_name))
    return targets


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in (
        "%B %d, %Y",
        "%b %d, %Y",
        "%d.%m.%Y",
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def infer_borrower_type(loan_type: str, current: str) -> Tuple[str, str]:
    norm = loan_type.casefold()
    if norm == "total loans":
        return "total", current
    if norm == "loans to individuals":
        return "individuals", "individuals"
    if norm == "loans to legal entities":
        return "legal_entities", "legal_entities"
    if current:
        return current, current
    return "needs_review", current


def parse_workbook(file_path: Path, report_period_folder: str, loaded_at: str) -> Tuple[List[dict], List[dict]]:
    wb = load_workbook(file_path, data_only=True)
    master_rows: List[dict] = []
    qa_rows: List[dict] = []

    for sheet in wb.worksheets:
        rows_read = 0
        output_rows = 0
        notes: List[str] = []
        status = "ok"

        period_b = parse_excel_date(sheet["B3"].value)
        period_c = parse_excel_date(sheet["C3"].value)
        if not period_b:
            notes.append("B3 period date not parsed")
        if not period_c:
            notes.append("C3 period date not parsed")
        unit_absolute = clean_text(sheet["D2"].value) or "billion UZS"

        max_row = sheet.max_row
        current_borrower = ""

        for row_idx in range(4, max_row + 1):
            loan_type = clean_text(sheet.cell(row=row_idx, column=1).value)
            if not loan_type:
                continue

            rows_read += 1
            borrower_type, current_borrower = infer_borrower_type(loan_type, current_borrower)

            for col_idx, metric_type in ((2, "absolute"), (3, "absolute"), (4, "growth")):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is None or clean_text(value) == "":
                    continue

                period_dt: Optional[date]
                if metric_type == "growth":
                    period_dt = period_c
                else:
                    period_dt = period_b if col_idx == 2 else period_c

                if period_dt:
                    period_year = period_dt.year
                    period_month = period_dt.month
                    period_date = period_dt.isoformat()
                    period_label = f"{period_dt.strftime('%B')} {period_dt.day}, {period_dt.year}"
                else:
                    period_info = parse_folder_period(report_period_folder)
                    if period_info:
                        period_year = period_info.year
                        period_month = period_info.month
                        period_date = period_info.dt.isoformat()
                        period_label = f"{period_info.dt.strftime('%B')} {period_info.dt.day}, {period_info.dt.year}"
                    else:
                        period_year = ""
                        period_month = ""
                        period_date = ""
                        period_label = ""

                unit = unit_absolute if metric_type == "absolute" else "percent"
                source_cell = f"{get_column_letter(col_idx)}{row_idx}"
                master_rows.append(
                    {
                        "report_period_folder": report_period_folder,
                        "period_year": period_year,
                        "period_month": period_month,
                        "period_date": period_date,
                        "period_label": period_label,
                        "loan_type": loan_type,
                        "borrower_type": borrower_type,
                        "metric_type": metric_type,
                        "value": value,
                        "unit": unit,
                        "source_file": str(file_path).replace("\\", "/"),
                        "source_sheet": sheet.title,
                        "source_cell": source_cell,
                        "loaded_at": loaded_at,
                    }
                )
                output_rows += 1

        if rows_read == 0:
            status = "warning"
            notes.append("No non-empty loan_type rows found in column A")
        if notes and status == "ok":
            status = "warning"

        qa_rows.append(
            {
                "source_file": str(file_path).replace("\\", "/"),
                "report_period_folder": report_period_folder,
                "source_sheet": sheet.title,
                "rows_read": rows_read,
                "output_rows_created": output_rows,
                "status": status,
                "notes": "; ".join(notes) if notes else "",
            }
        )

    wb.close()
    return master_rows, qa_rows


def write_csv(path: Path, columns: Iterable[str], rows: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(columns))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    ensure_outputs_writable(args.overwrite)

    targets = list_target_files(args.periods)
    if not targets:
        raise SystemExit("No matching files found for loan types parser.")

    loaded_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    all_master: List[dict] = []
    all_qa: List[dict] = []

    for file_path, folder in targets:
        master_rows, qa_rows = parse_workbook(file_path=file_path, report_period_folder=folder, loaded_at=loaded_at)
        all_master.extend(master_rows)
        all_qa.extend(qa_rows)

    write_csv(MASTER_OUT, MASTER_COLUMNS, all_master)
    write_csv(QA_OUT, QA_COLUMNS, all_qa)

    print(f"Wrote {len(all_master)} rows to {MASTER_OUT}")
    print(f"Wrote {len(all_qa)} rows to {QA_OUT}")


if __name__ == "__main__":
    main()
